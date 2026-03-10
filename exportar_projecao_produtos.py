#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exporta projeção de estoque por produto para Excel (mesma visão da tela Projeção de Estoque).

Usa a MESMA lógica do site: chama a API do dashboard quando disponível (variável DASHBOARD_URL
ou --api). Caso contrário usa SQL (pode divergir da tela).

Uso:
  python exportar_projecao_produtos.py "45.14.0035"
  python exportar_projecao_produtos.py "BRASIL TROPICAL, 45.14.0035"
  set DASHBOARD_URL=http://localhost:3000
  python exportar_projecao_produtos.py "45.14.0035" -o projecao.xlsx
"""

from __future__ import annotations

import os
import re
import sys
import urllib.request
import urllib.error
import urllib.parse
import json
from calendar import monthrange
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any

try:
    import pandas as pd
    import pyodbc
except ImportError as e:
    print(f"Erro: instale as dependências: pip install pandas pyodbc openpyxl\n{e}")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# --- Configuração (igual a outros scripts do projeto; pode sobrescrever via env) ---
DB_CONFIG = {
    "server": os.environ.get("DB_SERVER", "189.126.197.82"),
    "database": os.environ.get("DB_DATABASE", "LINX_PRODUCAO"),
    "username": os.environ.get("DB_USERNAME", "andre.nerd"),
    "password": os.environ.get("DB_PASSWORD", "nerd123@"),
}

# Filiais ScarfMe (inventory + e-commerce) - mesma lista do dashboard
SCARFME_FILIAIS_ESTOQUE = [
    "GUARULHOS - RSR", "IGUATEMI SP - JJJ", "MORUMBI - JJJ", "OSCAR FREIRE - FSZ",
    "SCARF ME - HIGIENOPOLIS 2", "SCARFME - IBIRAPUERA LLL", "SCARFME ME - PAULISTA FFF",
    "SCARF ME - PAULISTA RSR", "SCARF ME - MATRIZ", "SCARFME MATRIZ CMS",
    "SCARF ME - MATRIZ LLL", "SCARF ME MATRIZ - FFF", "VILLA LOBOS - LLL",
    "MSC COMERCIO DE LENCOS LT",
]
SCARFME_FILIAIS_VENDAS = [
    "GUARULHOS - RSR", "IGUATEMI SP - JJJ", "MORUMBI - JJJ", "OSCAR FREIRE - FSZ",
    "SCARF ME - HIGIENOPOLIS 2", "SCARFME - IBIRAPUERA LLL", "SCARFME ME - PAULISTA FFF",
    "SCARF ME - PAULISTA RSR", "SCARF ME - MATRIZ", "SCARFME MATRIZ CMS",
    "SCARF ME - MATRIZ LLL", "SCARF ME MATRIZ - FFF", "VILLA LOBOS - LLL",
    "MSC COMERCIO DE LENCOS LT",
]
# Varejo: vendas de loja sem e-commerce (igual backend quando "Todas as filiais")
SCARFME_FILIAIS_VAREJO = [
    f for f in SCARFME_FILIAIS_VENDAS if f not in (
        "SCARFME MATRIZ CMS", "SCARF ME - MATRIZ LLL", "SCARF ME MATRIZ - FFF", "MSC COMERCIO DE LENCOS LT"
    )
]
if not SCARFME_FILIAIS_VAREJO:
    SCARFME_FILIAIS_VAREJO = [
        "GUARULHOS - RSR", "IGUATEMI SP - JJJ", "MORUMBI - JJJ", "OSCAR FREIRE - FSZ",
        "SCARF ME - HIGIENOPOLIS 2", "SCARFME - IBIRAPUERA LLL", "SCARFME ME - PAULISTA FFF",
        "SCARF ME - PAULISTA RSR", "SCARF ME - MATRIZ", "VILLA LOBOS - LLL",
    ]
SCARFME_ECOM_FILIAIS = [
    "SCARFME MATRIZ CMS", "SCARF ME - MATRIZ LLL", "SCARF ME MATRIZ - FFF",
    "MSC COMERCIO DE LENCOS LT",
]
SCARFME_EXCLUDED_LINES = [
    "PRIVATE LABEL", "GASTRONOMICA", "PERFUMARIA", "CASHMERE",
    "ELETRONICOS", "EMBALAGENS", "CAPAS E ACESSORIOS P/ CEL",
]

MESES_NOMES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
MESES_NOMES_PT = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
FATOR_PROJECAO = 1.1  # vendas ano passado * 1.1 = projeção mensal


def fetch_projecao_via_api(base_url: str, company: str, produtos_set: set, filial: Optional[str] = None) -> Optional[List[Dict]]:
    """
    Chama a API do dashboard (mesma do site) e filtra por produto/código.
    Retorna lista no formato esperado por exportar_xlsx ou None se falhar.
    """
    base_url = base_url.rstrip("/")
    params = f"company={company}&dataType=projecao-mensal"
    if filial:
        params += f"&filial={urllib.parse.quote(filial)}"
    url = f"{base_url}/api/controle-estoque?{params}"
    try:
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode())
    except Exception as e:
        print(f"Aviso: API indisponível ({e}). Usando SQL.")
        return None
    raw = data.get("data") or data.get("projecao") or []
    if not isinstance(raw, list):
        return None
    resultados = []
    for cat in raw:
        produto = (cat.get("produto") or "").strip()
        descricao = (cat.get("descricao") or cat.get("categoria") or "").strip()
        if not produto and not descricao:
            continue
        # Filtrar: código exato ou nome contido
        match = False
        for p in produtos_set:
            p = p.strip().upper()
            if not p:
                continue
            if p in (produto.upper(), descricao.upper()):
                match = True
                break
            if produto.upper().find(p) >= 0 or descricao.upper().find(p) >= 0:
                match = True
                break
        if not match and produtos_set:
            continue
        meses = cat.get("meses") or []
        if len(meses) != 12:
            continue
        vendas_proj = []
        estoque_proj = []
        duracao_proj = []
        vendas_real = []
        for m in meses:
            vendas_proj.append(int(m.get("vendas") or 0))
            est = m.get("estoque")
            is_past = m.get("isMesPassado") is True
            if is_past:
                estoque_proj.append(None)
            elif est is not None:
                estoque_proj.append(int(est))
            else:
                estoque_proj.append(None)
            dur = m.get("duracao")
            duracao_proj.append(int(dur) if dur is not None and dur != 0 else None)
            vr = m.get("vendasReais")
            vendas_real.append(int(vr) if vr is not None else 0)
        # Estoque real só no mês atual
        mes_atual = datetime.now().month
        estoque_real = [None] * 12
        duracao_real = [None] * 12
        for i, m in enumerate(meses):
            if (i + 1) == mes_atual:
                if m.get("estoque") is not None:
                    estoque_real[i] = int(m.get("estoque") or 0)
                # duração real: backend não envia; calculamos se tivermos vendasReais e estoque
                if estoque_real[i] and vendas_real[i] and datetime.now().day > 0:
                    consumo = vendas_real[i] / datetime.now().day
                    if consumo > 0:
                        duracao_real[i] = round(estoque_real[i] / consumo)
                break
        resultados.append({
            "produto": produto,
            "descricao": descricao,
            "linha": (cat.get("linha") or "").strip(),
            "subgrupo": (cat.get("subgrupo") or "").strip(),
            "grade": (cat.get("grade") or "").strip(),
            "colecao": (cat.get("colecao") or "").strip(),
            "cor": (cat.get("cor") or "").strip() or "SEM COR",
            "vendas_proj": vendas_proj,
            "estoque_proj": estoque_proj,
            "duracao_proj": duracao_proj,
            "vendas_real": vendas_real,
            "estoque_real": estoque_real,
            "duracao_real": duracao_real,
        })
    return resultados if resultados else None


def normalizar(s: Optional[str]) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return str(s).strip().upper()


def parece_codigo(termo: str) -> bool:
    """True se o termo parece um código de produto (ex.: 45.14.0035)."""
    t = termo.strip()
    if not t:
        return False
    return bool(re.match(r"^[\d.\s]+$", t))


def conectar_banco():
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={DB_CONFIG['server']};"
        f"DATABASE={DB_CONFIG['database']};"
        f"UID={DB_CONFIG['username']};"
        f"PWD={DB_CONFIG['password']};"
    )
    return pyodbc.connect(conn_str, timeout=30)


def resolver_produtos(conn, termos: list[str]) -> list[str]:
    """
    Dado uma lista de termos (códigos ou nomes), retorna lista de PRODUTO (códigos)
    que atendem: código exato ou DESC_PRODUTO contendo o termo (case insensitive).
    """
    if not termos:
        return []
    codigos = []
    nomes = []
    for t in termos:
        t = t.strip()
        if not t:
            continue
        if parece_codigo(t):
            codigos.append(t.strip())
        else:
            nomes.append(t.strip())

    placeholders_cod = ", ".join(["?"] * len(codigos)) if codigos else ""
    placeholders_nome = " OR ".join(
        ["UPPER(LTRIM(RTRIM(ISNULL(p.DESC_PRODUTO, '')))) LIKE ?"] * len(nomes)
    ) if nomes else ""

    condicoes = []
    params = []
    if placeholders_cod:
        condicoes.append(f"p.PRODUTO IN ({placeholders_cod})")
        params.extend(codigos)
    if placeholders_nome:
        condicoes.append(f"({placeholders_nome})")
        for n in nomes:
            params.append(f"%{n.upper()}%")

    if not condicoes:
        return []

    where = " OR ".join(condicoes)
    sql = f"""
    SELECT DISTINCT p.PRODUTO
    FROM PRODUTOS p WITH (NOLOCK)
    WHERE {where}
      AND UPPER(LTRIM(RTRIM(ISNULL(p.LINHA, '')))) NOT IN ({','.join(['?']*len(SCARFME_EXCLUDED_LINES))})
    """
    params.extend(SCARFME_EXCLUDED_LINES)
    df = pd.read_sql(sql, conn, params=params)
    return df["PRODUTO"].astype(str).str.strip().unique().tolist()


def buscar_estoque_por_produto_cor(conn, produtos: list[str]) -> pd.DataFrame:
    """Estoque atual por PRODUTO + COR (agregado nas filiais ScarfMe)."""
    if not produtos:
        return pd.DataFrame()
    ph = ", ".join(["?"] * len(produtos))
    ph_fil = ", ".join(["?"] * len(SCARFME_FILIAIS_ESTOQUE))
    params = list(SCARFME_FILIAIS_ESTOQUE) + list(produtos)
    sql = f"""
    SELECT
        p.PRODUTO,
        UPPER(LTRIM(RTRIM(ISNULL(p.DESC_PRODUTO, '')))) AS descricao,
        UPPER(LTRIM(RTRIM(ISNULL(p.LINHA, '')))) AS linha,
        UPPER(LTRIM(RTRIM(ISNULL(p.SUBGRUPO_PRODUTO, '')))) AS subgrupo,
        UPPER(LTRIM(RTRIM(ISNULL(CONVERT(VARCHAR, p.GRADE), '')))) AS grade,
        UPPER(LTRIM(RTRIM(ISNULL(p.COLECAO, '')))) AS colecao,
        UPPER(LTRIM(RTRIM(ISNULL(e.COR_PRODUTO, '')))) AS cor,
        SUM(CASE WHEN e.ESTOQUE > 0 THEN e.ESTOQUE ELSE 0 END) AS estoque
    FROM ESTOQUE_PRODUTOS e WITH (NOLOCK)
    LEFT JOIN PRODUTOS p WITH (NOLOCK) ON e.PRODUTO = p.PRODUTO
    WHERE e.FILIAL IN ({ph_fil})
      AND e.PRODUTO IN ({ph})
      AND e.ESTOQUE > 0
      AND UPPER(LTRIM(RTRIM(ISNULL(p.LINHA, '')))) NOT IN ({','.join(['?']*len(SCARFME_EXCLUDED_LINES))})
    GROUP BY p.PRODUTO, p.DESC_PRODUTO, p.LINHA, p.SUBGRUPO_PRODUTO, p.GRADE, p.COLECAO, e.COR_PRODUTO
    """
    params.extend(SCARFME_EXCLUDED_LINES)
    return pd.read_sql(sql, conn, params=params)


def buscar_vendas_ano_por_mes(conn, produtos: list[str], ano: int) -> pd.DataFrame:
    """Vendas varejo por PRODUTO, COR e mês. Backend usa vp.FILIAL IN (...). Tenta FILIAL; se falhar, usa JOIN em FILIAIS."""
    if not produtos:
        return pd.DataFrame()
    ph = ", ".join(["?"] * len(produtos))
    ph_fil = ", ".join(["?"] * len(SCARFME_FILIAIS_VAREJO))
    params = [ano] + list(SCARFME_FILIAIS_VAREJO) + list(produtos)
    sql = f"""
    SELECT
        vp.PRODUTO,
        UPPER(LTRIM(RTRIM(ISNULL(vp.COR_PRODUTO, '')))) AS cor,
        MONTH(vp.DATA_VENDA) AS mes,
        SUM(CASE WHEN vp.QTDE_CANCELADA = 0 THEN vp.QTDE ELSE 0 END) AS vendas
    FROM W_CTB_LOJA_VENDA_PEDIDO_PRODUTO vp WITH (NOLOCK)
    WHERE YEAR(vp.DATA_VENDA) = ?
      AND vp.FILIAL IN ({ph_fil})
      AND vp.PRODUTO IN ({ph})
      AND vp.QTDE > 0
    GROUP BY vp.PRODUTO, vp.COR_PRODUTO, MONTH(vp.DATA_VENDA)
    """
    try:
        df = pd.read_sql(sql, conn, params=params)
        if not df.empty:
            return df
    except Exception:
        pass
    sql_join = f"""
    SELECT
        vp.PRODUTO,
        UPPER(LTRIM(RTRIM(ISNULL(vp.COR_PRODUTO, '')))) AS cor,
        MONTH(vp.DATA_VENDA) AS mes,
        SUM(CASE WHEN vp.QTDE_CANCELADA = 0 THEN vp.QTDE ELSE 0 END) AS vendas
    FROM W_CTB_LOJA_VENDA_PEDIDO_PRODUTO vp WITH (NOLOCK)
    INNER JOIN FILIAIS f WITH (NOLOCK) ON f.COD_FILIAL = vp.CODIGO_FILIAL
    WHERE YEAR(vp.DATA_VENDA) = ?
      AND LTRIM(RTRIM(f.FILIAL)) IN ({ph_fil})
      AND vp.PRODUTO IN ({ph})
      AND vp.QTDE > 0
    GROUP BY vp.PRODUTO, vp.COR_PRODUTO, MONTH(vp.DATA_VENDA)
    """
    return pd.read_sql(sql_join, conn, params=params)


def buscar_vendas_ecommerce_ano_por_mes(conn, produtos: list[str], ano: int) -> pd.DataFrame:
    """Vendas e-commerce por PRODUTO, COR e mês (ano dado)."""
    if not produtos:
        return pd.DataFrame()
    ph = ", ".join(["?"] * len(produtos))
    ph_fil = ", ".join(["?"] * len(SCARFME_ECOM_FILIAIS))
    params = [ano] + list(SCARFME_ECOM_FILIAIS) + list(produtos)
    data_ec = "COALESCE(CAST(fp.ENTREGA AS DATE), CAST(f.EMISSAO AS DATE))"
    sql = f"""
    SELECT
        fp.PRODUTO,
        UPPER(LTRIM(RTRIM(ISNULL(fp.COR_PRODUTO, '')))) AS cor,
        MONTH({data_ec}) AS mes,
        SUM(CAST(fp.QTDE AS FLOAT)) AS vendas
    FROM FATURAMENTO f WITH (NOLOCK)
    JOIN W_FATURAMENTO_PROD_02 fp WITH (NOLOCK)
      ON f.FILIAL = fp.FILIAL AND f.NF_SAIDA = fp.NF_SAIDA AND f.SERIE_NF = fp.SERIE_NF
    WHERE YEAR({data_ec}) = ?
      AND f.NOTA_CANCELADA = 0
      AND f.NATUREZA_SAIDA IN ('100.02', '100.022')
      AND CAST(fp.QTDE AS FLOAT) > 0
      AND REPLACE(REPLACE(LTRIM(RTRIM(f.FILIAL)), NCHAR(0x00A0), ' '), CHAR(9), ' ') IN ({ph_fil})
      AND fp.PRODUTO IN ({ph})
    GROUP BY fp.PRODUTO, fp.COR_PRODUTO, MONTH({data_ec})
    """
    return pd.read_sql(sql, conn, params=params)


def buscar_vendas_mes_atual(conn, produtos: list[str], inicio_mes, fim_mes) -> pd.DataFrame:
    """Vendas (varejo + e-commerce) do mês atual até hoje. Varejo: vp.FILIAL IN (filiais varejo)."""
    if not produtos:
        return pd.DataFrame()
    ph = ", ".join(["?"] * len(produtos))
    ph_fil = ", ".join(["?"] * len(SCARFME_FILIAIS_VAREJO))
    params = [inicio_mes, fim_mes] + list(SCARFME_FILIAIS_VAREJO) + list(produtos)
    sql = f"""
    SELECT
        vp.PRODUTO,
        UPPER(LTRIM(RTRIM(ISNULL(vp.COR_PRODUTO, '')))) AS cor,
        SUM(CASE WHEN vp.QTDE_CANCELADA = 0 THEN vp.QTDE ELSE 0 END) AS vendas
    FROM W_CTB_LOJA_VENDA_PEDIDO_PRODUTO vp WITH (NOLOCK)
    WHERE vp.DATA_VENDA >= ? AND vp.DATA_VENDA < ?
      AND vp.FILIAL IN ({ph_fil})
      AND vp.PRODUTO IN ({ph})
      AND vp.QTDE > 0
    GROUP BY vp.PRODUTO, vp.COR_PRODUTO
    """
    try:
        df_v = pd.read_sql(sql, conn, params=params)
    except Exception:
        ph_fil = ", ".join(["?"] * len(SCARFME_FILIAIS_VAREJO))
        params = [inicio_mes, fim_mes] + list(SCARFME_FILIAIS_VAREJO) + list(produtos)
        sql = f"""
    SELECT
        vp.PRODUTO,
        UPPER(LTRIM(RTRIM(ISNULL(vp.COR_PRODUTO, '')))) AS cor,
        SUM(CASE WHEN vp.QTDE_CANCELADA = 0 THEN vp.QTDE ELSE 0 END) AS vendas
    FROM W_CTB_LOJA_VENDA_PEDIDO_PRODUTO vp WITH (NOLOCK)
    INNER JOIN FILIAIS f WITH (NOLOCK) ON f.COD_FILIAL = vp.CODIGO_FILIAL
    WHERE vp.DATA_VENDA >= ? AND vp.DATA_VENDA < ?
      AND LTRIM(RTRIM(f.FILIAL)) IN ({ph_fil})
      AND vp.PRODUTO IN ({ph})
      AND vp.QTDE > 0
    GROUP BY vp.PRODUTO, vp.COR_PRODUTO
    """
        df_v = pd.read_sql(sql, conn, params=params)
    ph_fil_ec = ", ".join(["?"] * len(SCARFME_ECOM_FILIAIS))
    params_ec = [inicio_mes, fim_mes] + list(SCARFME_ECOM_FILIAIS) + list(produtos)
    sql_ec = f"""
    SELECT
        fp.PRODUTO,
        UPPER(LTRIM(RTRIM(ISNULL(fp.COR_PRODUTO, '')))) AS cor,
        SUM(CAST(fp.QTDE AS FLOAT)) AS vendas
    FROM FATURAMENTO f WITH (NOLOCK)
    JOIN W_FATURAMENTO_PROD_02 fp WITH (NOLOCK)
      ON f.FILIAL = fp.FILIAL AND f.NF_SAIDA = fp.NF_SAIDA AND f.SERIE_NF = fp.SERIE_NF
    WHERE f.EMISSAO >= ? AND f.EMISSAO < ?
      AND f.NOTA_CANCELADA = 0
      AND f.NATUREZA_SAIDA IN ('100.02', '100.022')
      AND CAST(fp.QTDE AS FLOAT) > 0
      AND REPLACE(REPLACE(LTRIM(RTRIM(f.FILIAL)), NCHAR(0x00A0), ' '), CHAR(9), ' ') IN ({ph_fil_ec})
      AND fp.PRODUTO IN ({ph})
    GROUP BY fp.PRODUTO, fp.COR_PRODUTO
    """
    df_ec = pd.read_sql(sql_ec, conn, params=params_ec)
    if df_ec.empty:
        return df_v
    if df_v.empty:
        return df_ec
    merged = df_v.merge(
        df_ec, on=["PRODUTO", "cor"], how="outer", suffixes=("_v", "_ec")
    ).fillna(0)
    merged["vendas"] = merged["vendas_v"] + merged["vendas_ec"]
    return merged[["PRODUTO", "cor", "vendas"]]


def build_projecao_mensal(
    estoque_df: pd.DataFrame,
    vendas_ano_passado_df: pd.DataFrame,
    vendas_mes_atual_series: dict,
    vendas_ano_atual_df: pd.DataFrame,
    ano_atual: int,
    mes_atual: int,
) -> list[dict]:
    """
    Replica a lógica do backend (fetchProjecaoMensal):
    - Vendas projeção = (vendas ano passado no mês) × 1,1.
    - Estoque projeção = estoque no INÍCIO de cada mês; depois descontar vendas (no mês atual: proj − real).
    - Duração = segundo passe: a partir do estoque no início do mês, quantos dias até zerar.
    """
    hoje = datetime.now()
    dias_corridos = hoje.day
    dias_no_mes_atual = monthrange(ano_atual, mes_atual)[1]
    dias_restantes_mes_atual = max(0, dias_no_mes_atual - dias_corridos)
    resultados = []

    for _, row in estoque_df.iterrows():
        produto = str(row["PRODUTO"]).strip()
        cor = normalizar(row["cor"])
        key = (produto, cor)
        estoque_inicial = int(row["estoque"] or 0)

        vendas_ap = vendas_ano_passado_df[
            (vendas_ano_passado_df["PRODUTO"] == produto)
            & (vendas_ano_passado_df["cor"].astype(str).str.strip().str.upper() == cor)
        ]
        vendas_por_mes_ap = vendas_ap.set_index("mes")["vendas"].to_dict() if not vendas_ap.empty else {}
        total_ap = sum(vendas_por_mes_ap.values()) or 0
        projecao_mensal_media = (total_ap / 12 * FATOR_PROJECAO) if total_ap > 0 else 0

        vendas_atual = vendas_ano_atual_df[
            (vendas_ano_atual_df["PRODUTO"] == produto)
            & (vendas_ano_atual_df["cor"].astype(str).str.strip().str.upper() == cor)
        ]
        vendas_real_por_mes = vendas_atual.set_index("mes")["vendas"].to_dict() if not vendas_atual.empty else {}
        vendas_mes_atual_val = int(vendas_mes_atual_series.get(key, 0) or 0)

        vendas_proj = [0] * 12
        estoque_proj = [None] * 12
        vendas_real_arr = [0] * 12
        est = estoque_inicial

        for i in range(12):
            mes_num = i + 1
            is_mes_atual = mes_num == mes_atual
            is_mes_passado = mes_num < mes_atual

            v_ap = vendas_por_mes_ap.get(mes_num, 0) or 0
            if v_ap > 0:
                vendas_proj[i] = max(0, int(round(v_ap * FATOR_PROJECAO)))
            else:
                vendas_proj[i] = max(0, int(round(projecao_mensal_media)))

            if is_mes_passado:
                vendas_real_arr[i] = int(vendas_real_por_mes.get(mes_num, 0) or 0)
                estoque_proj[i] = None
                continue

            estoque_no_inicio_do_mes = est
            estoque_proj[i] = estoque_no_inicio_do_mes

            if is_mes_atual:
                vendas_real_arr[i] = vendas_mes_atual_val
                vendas_a_descontar = max(0, vendas_proj[i] - vendas_mes_atual_val)
            else:
                vendas_real_arr[i] = int(vendas_real_por_mes.get(mes_num, 0) or 0)
                vendas_a_descontar = vendas_proj[i]

            est = max(0, est - vendas_a_descontar)

        duracao_proj = [None] * 12
        for i in range(12):
            if estoque_proj[i] is None or estoque_proj[i] <= 0:
                continue
            remaining = estoque_proj[i]
            total_dias = 0
            ultimo_consumo_diario = 0.0
            for j in range(i, 12):
                mes_num_j = j + 1
                is_atual_j = mes_num_j == mes_atual
                vendas_mes_j = vendas_proj[j]
                if is_atual_j and vendas_real_arr[j] is not None:
                    vendas_mes_j = max(0, vendas_proj[j] - vendas_real_arr[j])
                dias_no_mes_j = dias_restantes_mes_atual if (is_atual_j and dias_restantes_mes_atual > 0) else monthrange(ano_atual, mes_num_j)[1]
                if dias_no_mes_j <= 0 or vendas_mes_j <= 0:
                    continue
                consumo_diario = vendas_mes_j / dias_no_mes_j
                ultimo_consumo_diario = consumo_diario
                dias_para_esvaziar = remaining / consumo_diario
                if dias_para_esvaziar >= dias_no_mes_j:
                    total_dias += dias_no_mes_j
                    remaining -= vendas_mes_j
                else:
                    total_dias += round(dias_para_esvaziar)
                    remaining = 0
                    break
            if remaining > 0 and ultimo_consumo_diario > 0:
                total_dias += round(remaining / ultimo_consumo_diario)
            elif remaining > 0 and i < 12:
                ultimo_mes_num = 12
                dias_ultimo = monthrange(ano_atual, ultimo_mes_num)[1]
                if dias_ultimo > 0 and vendas_proj[11] > 0:
                    total_dias += round(remaining / (vendas_proj[11] / dias_ultimo))
            duracao_proj[i] = int(total_dias) if total_dias > 0 else 0

        estoque_real_mes_atual = estoque_inicial
        if dias_corridos > 0 and vendas_mes_atual_val > 0 and estoque_real_mes_atual > 0:
            duracao_real_mes_atual = round(estoque_real_mes_atual / (vendas_mes_atual_val / dias_corridos))
        else:
            duracao_real_mes_atual = 0

        resultados.append({
            "produto": produto,
            "descricao": row.get("descricao", ""),
            "linha": row.get("linha", ""),
            "subgrupo": row.get("subgrupo", ""),
            "grade": row.get("grade", ""),
            "colecao": row.get("colecao", ""),
            "cor": cor,
            "vendas_proj": vendas_proj,
            "estoque_proj": estoque_proj,
            "duracao_proj": duracao_proj,
            "vendas_real": vendas_real_arr,
            "estoque_real": [
                estoque_real_mes_atual if (i + 1) == mes_atual else None for i in range(12)
            ],
            "duracao_real": [
                duracao_real_mes_atual if (i + 1) == mes_atual else None for i in range(12)
            ],
        })
    return resultados


def exportar_xlsx(resultados: list[dict], caminho: str) -> None:
    """Gera o arquivo Excel com a mesma visão da tela: um bloco por item (categoria + 6 linhas de tipo)."""
    if not resultados:
        return

    if openpyxl is None:
        linhas = []
        for r in resultados:
            cat = f"{r['descricao']} | {r['produto']} | {r['linha']} | {r['subgrupo']} | {r['grade']} | {r['colecao']} | {r['cor']}"
            for tipo, vals in [
                ("VENDA (PROJEÇÃO)", r["vendas_proj"]),
                ("ESTOQUE (PROJEÇÃO)", r["estoque_proj"]),
                ("DURACAO (PROJEÇÃO)", r["duracao_proj"]),
                ("VENDA (REAL)", r["vendas_real"]),
                ("ESTOQUE (REAL)", r["estoque_real"]),
                ("DURACAO (REAL)", r["duracao_real"]),
            ]:
                linha = [cat, tipo]
                for v in vals:
                    if v is None:
                        linha.append("")
                    elif isinstance(v, int) and tipo.startswith("DURACAO") and v > 0:
                        linha.append(f"{v} dias")
                    else:
                        linha.append(v)
                linhas.append(linha)
        cols = ["CATEGORIA", "TIPO"] + MESES_NOMES
        df = pd.DataFrame(linhas, columns=cols)
        df.to_excel(caminho, index=False, sheet_name="Projeção")
        print(f"Arquivo salvo (pandas): {caminho}")
        return

    wb = openpyxl.Workbook()
    ws_regras = wb.active
    ws_regras.title = "Regras de projeção"
    _escrever_sheet_regras(ws_regras, datetime.now())
    ws = wb.create_sheet("Projeção", 0)

    hoje = datetime.now()
    mes_atual = hoje.month
    thin = Side(style="thin")

    # Linha 1: Título
    ws.merge_cells("A1:N1")
    titulo = ws["A1"]
    titulo.value = f"Projeção de Estoque por Produto — Gerado em {hoje:%d/%m/%Y %H:%M}"
    titulo.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    titulo.font = Font(bold=True, size=12, color="FFFFFF")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Linha 2: Cabeçalho da tabela
    header = ["CATEGORIA", "TIPO"] + MESES_NOMES
    for col, val in enumerate(header, 1):
        c = ws.cell(row=2, column=col, value=val)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        c.border = Border(
            left=thin, right=thin, top=thin, bottom=Side(style="medium"),
        )
    ws.row_dimensions[2].height = 24

    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_proj = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    fill_real = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    fill_categoria = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_tipo = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_duracao_alerta = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    fill_mes_atual = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    row_num = 3
    for idx, r in enumerate(resultados):
        cat_label = (
            f"{r['descricao'] or r['produto']}  |  Cód: {r['produto']}  |  "
            f"Linha: {r['linha']}  |  Subgrupo: {r['subgrupo']}  |  Grade: {r['grade']}  |  "
            f"Coleção: {r['colecao']}  |  {r['cor'] or 'SEM COR'}"
        )
        blocos = [
            ("VENDA (PROJEÇÃO)", r["vendas_proj"], "proj", False),
            ("ESTOQUE (PROJEÇÃO)", r["estoque_proj"], "proj", False),
            ("DURACAO (PROJEÇÃO)", r["duracao_proj"], "proj", True),
            ("VENDA (REAL)", r["vendas_real"], "real", False),
            ("ESTOQUE (REAL)", r["estoque_real"], "real", False),
            ("DURACAO (REAL)", r["duracao_real"], "real", True),
        ]
        row_inicio_bloco = row_num
        for tipo, vals, tipo_fill, is_duracao in blocos:
            ws.cell(row=row_num, column=1, value=cat_label)
            ws.cell(row=row_num, column=2, value=tipo)
            ws.cell(row=row_num, column=1).fill = fill_categoria
            ws.cell(row=row_num, column=1).border = border_cell
            ws.cell(row=row_num, column=2).fill = fill_tipo if tipo_fill == "proj" else fill_real
            ws.cell(row=row_num, column=2).border = border_cell
            ws.cell(row=row_num, column=2).font = Font(size=10, bold=(tipo_fill == "proj"))
            for i, v in enumerate(vals):
                col = 3 + i
                is_mes_atual = (i + 1) == mes_atual
                if v is None:
                    cell_val = ""
                elif is_duracao and isinstance(v, (int, float)) and v > 0:
                    cell_val = f"{int(v)} dias"
                else:
                    cell_val = v
                cell = ws.cell(row=row_num, column=col, value=cell_val)
                cell.border = border_cell
                cell.alignment = Alignment(horizontal="center" if col > 2 else "left", vertical="center")
                if tipo_fill == "real":
                    cell.fill = fill_real
                else:
                    cell.fill = fill_proj
                if is_mes_atual:
                    cell.fill = fill_mes_atual
                if is_duracao and isinstance(v, (int, float)) and 0 < v <= 120:
                    cell.fill = fill_duracao_alerta
                    cell.font = Font(bold=True, color="B91C1C")
            row_num += 1
        if row_num - 1 > row_inicio_bloco:
            ws.merge_cells(start_row=row_inicio_bloco, start_column=1, end_row=row_num - 1, end_column=1)
            ws.cell(row=row_inicio_bloco, column=1).alignment = Alignment(vertical="center", wrap_text=True)
        row_num += 1

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 22
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.freeze_panes = "C3"
    wb.save(caminho)
    print(f"Arquivo salvo: {caminho}")


def exportar_txt(resultados: List[Dict], caminho_txt: str, mes_atual: int, ano_atual: int) -> None:
    """
    Gera um TXT com: período (mês atual), para cada produto código, descrição,
    total vendas (mês atual), duração real ou (se vendas 0) duração projetada.
    """
    idx_mes = mes_atual - 1
    periodo_nome = MESES_NOMES_PT[idx_mes] if 0 <= idx_mes < len(MESES_NOMES_PT) else str(mes_atual)
    periodo_str = f"{periodo_nome}/{ano_atual}"
    linhas = [
        f"Período (mês atual): {periodo_str}",
        "-----------------------------------------------",
        "",
    ]
    for r in resultados:
        codigo = (r.get("produto") or "").strip()
        descricao = (r.get("descricao") or "").strip()
        vendas_list = r.get("vendas_real") or [0] * 12
        total_vendas = int(vendas_list[idx_mes]) if idx_mes < len(vendas_list) else 0
        duracao_real_list = r.get("duracao_real") or [None] * 12
        duracao_proj_list = r.get("duracao_proj") or [None] * 12
        duracao_real = duracao_real_list[idx_mes] if idx_mes < len(duracao_real_list) else None
        duracao_proj = duracao_proj_list[idx_mes] if idx_mes < len(duracao_proj_list) else None
        if total_vendas > 0 and duracao_real is not None:
            duracao_exibir = f"{int(duracao_real)} dias (real)"
        else:
            duracao_exibir = f"{int(duracao_proj)} dias (projetada)" if duracao_proj is not None else "-"
        linhas.extend([codigo, descricao, str(total_vendas), duracao_exibir, ""])
    with open(caminho_txt, "w", encoding="utf-8") as f:
        f.write("\n".join(linhas))
    print(f"Resumo TXT salvo: {caminho_txt}")


def _escrever_sheet_regras(ws, hoje: datetime) -> None:
    """Preenche a aba 'Regras de projeção' com as regras utilizadas no cálculo."""
    ws.column_dimensions["A"].width = 90
    titulo = ws["A1"]
    titulo.value = "Regras de projeção de estoque"
    titulo.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[1].height = 26
    regras = [
        "",
        "1. VENDA (PROJEÇÃO)",
        "   • Por mês: (Vendas do mesmo mês no ano anterior) × 1,1",
        "   • Quando não há vendas no ano anterior no mês, a projeção do mês é 0.",
        "",
        "2. ESTOQUE (PROJEÇÃO)",
        "   • Estoque inicial = estoque atual (soma de todas as filiais consideradas).",
        "   • Para o mês atual: descontar do estoque a projeção mensal menos as vendas reais já realizadas no mês.",
        "   • Para os meses seguintes: descontar a projeção de vendas do mês do estoque do mês anterior.",
        "   • Estoque projetado = máximo(0, estoque anterior − vendas a descontar).",
        "",
        "3. DURACAO (PROJEÇÃO) — em dias",
        "   • Duração = (Estoque projetado no mês ÷ Vendas projetadas do mês) × número de dias do mês.",
        "   • Se vendas projetadas = 0 e estoque > 0: exibe 999 dias.",
        "   • Valores em vermelho: duração ≤ 120 dias (alerta de estoque baixo).",
        "",
        "4. VENDA (REAL)",
        "   • Quantidade realmente vendida no mês (varejo + e-commerce), por mês do ano atual.",
        "",
        "5. ESTOQUE (REAL)",
        "   • Estoque atual (apenas no mês corrente; meses passados/futuros em branco).",
        "",
        "6. DURACAO (REAL) — em dias",
        "   • Com o estoque atual e o ritmo de venda do mês até hoje:",
        "   • Consumo diário = Vendas no mês até hoje ÷ Dias corridos.",
        "   • Duração = Estoque atual ÷ Consumo diário.",
        "",
        f"Relatório gerado em: {hoje:%d/%m/%Y %H:%M}",
    ]
    for i, texto in enumerate(regras, start=2):
        cell = ws.cell(row=i, column=1, value=texto)
        if any(texto.strip().startswith(f"{n}.") for n in range(1, 7)):
            cell.font = Font(bold=True, size=11)
        elif texto.startswith("   "):
            cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 18 if texto.startswith("   ") else 22


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Exporta projeção de estoque por produto para Excel (filtro por código ou nome, vários separados por vírgula)."
    )
    parser.add_argument(
        "produtos",
        type=str,
        nargs="?",
        default=None,
        help='Códigos ou nomes de produtos separados por vírgula (ex.: "45.14.0035" ou "BRASIL TROPICAL, LENÇOS"). Se omitir, será solicitado.',
    )
    parser.add_argument(
        "--saida", "-o",
        default=None,
        help="Caminho do arquivo Excel de saída (default: projecao_produtos_YYYY-MM-DD.xlsx)",
    )
    parser.add_argument(
        "--api",
        default=None,
        help="URL base do dashboard para usar a API (ex.: http://localhost:3000). Se não informar, usa variável DASHBOARD_URL. Com API, a projeção usa exatamente a mesma lógica do site.",
    )
    args = parser.parse_args()

    termos = [t.strip() for t in args.produtos.split(",") if t.strip()] if args.produtos else []
    if not termos:
        try:
            args.produtos = input("Digite os códigos ou nomes dos produtos (separados por vírgula): ").strip()
        except EOFError:
            args.produtos = ""
        termos = [t.strip() for t in args.produtos.split(",") if t.strip()]
    if not termos:
        print("Nenhum produto informado. Exemplo: python exportar_projecao_produtos.py \"45.14.0035\"")
        sys.exit(1)

    produtos_set = {t.upper() for t in termos}
    base_url = args.api or os.environ.get("DASHBOARD_URL")
    resultados = None
    if base_url:
        print(f"Buscando projeção via API do dashboard ({base_url})...")
        resultados = fetch_projecao_via_api(base_url.strip(), "scarfme", produtos_set)
        if resultados:
            print(f"Encontrados {len(resultados)} item(ns) na projeção (API).")

    if resultados is None:
        # Fallback: SQL (pode divergir da tela)
        hoje = datetime.now()
        ano_atual = hoje.year
        mes_atual = hoje.month
        inicio_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        fim_mes = hoje + timedelta(days=1)  # exclusivo: inclui todo o dia de hoje

        print("Conectando ao banco...")
        conn = conectar_banco()
        try:
            print("Resolvendo produtos...")
            produtos = resolver_produtos(conn, termos)
            if not produtos:
                print("Nenhum produto encontrado para os termos informados.")
                sys.exit(1)
            print(f"Encontrados {len(produtos)} produto(s): {', '.join(produtos[:10])}{'...' if len(produtos) > 10 else ''}")

            print("Buscando estoque por produto/cor...")
            estoque_df = buscar_estoque_por_produto_cor(conn, produtos)
            if estoque_df.empty:
                print("Nenhum estoque encontrado para esses produtos.")
                sys.exit(1)
            print(f"Linhas (produto+cor) com estoque: {len(estoque_df)}")

            print("Buscando vendas ano passado...")
            vendas_ap = buscar_vendas_ano_por_mes(conn, produtos, ano_atual - 1)
            vendas_ap_ec = buscar_vendas_ecommerce_ano_por_mes(conn, produtos, ano_atual - 1)
            if not vendas_ap_ec.empty:
                if vendas_ap.empty:
                    vendas_ap = vendas_ap_ec.copy()
                else:
                    vendas_ap = pd.concat([
                        vendas_ap,
                        vendas_ap_ec,
                    ]).groupby(["PRODUTO", "cor", "mes"], as_index=False).agg({"vendas": "sum"})
            if "cor" not in vendas_ap.columns:
                vendas_ap["cor"] = ""

            print("Buscando vendas mês atual...")
            vendas_mes_atual_df = buscar_vendas_mes_atual(conn, produtos, inicio_mes, fim_mes)
            vendas_mes_atual_series = {}
            if not vendas_mes_atual_df.empty:
                for _, r in vendas_mes_atual_df.iterrows():
                    vendas_mes_atual_series[(str(r["PRODUTO"]).strip(), normalizar(r["cor"]))] = int(r["vendas"] or 0)

            print("Buscando vendas ano atual (real)...")
            vendas_atual = buscar_vendas_ano_por_mes(conn, produtos, ano_atual)
            vendas_atual_ec = buscar_vendas_ecommerce_ano_por_mes(conn, produtos, ano_atual)
            if not vendas_atual_ec.empty:
                if vendas_atual.empty:
                    vendas_atual = vendas_atual_ec.copy()
                else:
                    vendas_atual = pd.concat([
                        vendas_atual,
                        vendas_atual_ec,
                    ]).groupby(["PRODUTO", "cor", "mes"], as_index=False).agg({"vendas": "sum"})
            if vendas_atual.empty:
                vendas_atual = pd.DataFrame(columns=["PRODUTO", "cor", "mes", "vendas"])
            if "cor" not in vendas_atual.columns:
                vendas_atual["cor"] = ""

            print("Calculando projeção (SQL)...")
            resultados = build_projecao_mensal(
                estoque_df,
                vendas_ap,
                vendas_mes_atual_series,
                vendas_atual,
                ano_atual,
                mes_atual,
            )
        finally:
            conn.close()

    if not resultados:
        print("Nenhum dado de projeção encontrado para os filtros informados.")
        sys.exit(1)

    hoje = datetime.now()
    saida = args.saida or f"projecao_produtos_{hoje:%Y-%m-%d}.xlsx"
    exportar_xlsx(resultados, saida)
    # TXT com período, código, descrição, total vendas e duração (real ou projetada)
    path_txt = os.path.splitext(saida)[0] + ".txt"
    exportar_txt(resultados, path_txt, hoje.month, hoje.year)
    print(f"Pronto. {len(resultados)} item(ns) exportado(s) para {saida}.")


if __name__ == "__main__":
    main()
