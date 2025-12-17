#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para exportar dados de clientes e vendas para Excel
Baseado na lógica da sessão de clientes do dashboard
"""

import os
import sys
import argparse
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================
# CONFIGURAÇÕES DO BANCO DE DADOS
# ============================================
DB_SERVER = '177.92.78.250'
DB_DATABASE = 'LINX_PRODUCAO'
DB_USERNAME = 'andre.nerd'
DB_PASSWORD = 'nerd123@'
DB_PORT = '1433'

# Configuração de empresas (mesma lógica do TypeScript)
COMPANIES = {
    'nerd': {
        'name': 'NERD',
        'filiais': [
            'NERD CENTER NORTE',
            'NERD HIGIENOPOLIS',
            'NERD LEBLON',
            'NERD MORUMBI RDRRRJ',
            'NERD VILLA LOBOS',
        ],
        'filial_display_names': {
            'NERD CENTER NORTE': 'CENTER NORTE',
            'NERD HIGIENOPOLIS': 'HIGIENOPOLIS',
            'NERD LEBLON': 'LEBLON',
            'NERD MORUMBI RDRRRJ': 'MORUMBI',
            'NERD VILLA LOBOS': 'VILLA LOBOS',
        }
    },
    'scarfme': {
        'name': 'SCARF ME',
        'filiais': [
            'GUARULHOS - RSR',
            'IGUATEMI SP - JJJ',
            'MORUMBI - JJJ',
            'OSCAR FREIRE - FSZ',
            'SCARF ME - HIGIENOPOLIS 2',
            'SCARFME - IBIRAPUERA LLL',
            'SCARFME ME - PAULISTA FFF',
            'SCARF ME - MATRIZ',
            'SCARFME MATRIZ CMS',
            'VILLA LOBOS - LLL',
        ],
        'filial_display_names': {
            'GUARULHOS - RSR': 'GUARULHOS',
            'IGUATEMI SP - JJJ': 'IGUATEMI',
            'MORUMBI - JJJ': 'MORUMBI',
            'OSCAR FREIRE - FSZ': 'OSCAR FREIRE',
            'SCARF ME - HIGIENOPOLIS 2': 'HIGIENÓPOLIS',
            'SCARFME - IBIRAPUERA LLL': 'IBIRAPUERA',
            'SCARFME ME - PAULISTA FFF': 'PAULISTA',
            'SCARF ME - MATRIZ': 'MATRIZ',
            'SCARFME MATRIZ CMS': 'E-COMMERCE',
            'VILLA LOBOS - LLL': 'VILLA LOBOS',
        },
        'ecommerce_filiais': ['SCARFME MATRIZ CMS']
    }
}

def get_db_connection():
    """Cria conexão com o banco de dados SQL Server"""
    connection_string = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={DB_SERVER},{DB_PORT};"
        f"DATABASE={DB_DATABASE};"
        f"UID={DB_USERNAME};"
        f"PWD={DB_PASSWORD};"
        f"TrustServerCertificate=yes;"
    )
    
    try:
        conn = pyodbc.connect(connection_string, timeout=60)
        return conn
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        raise

def parse_date(date_str: str) -> date:
    """Converte string de data para objeto date"""
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        raise ValueError(f"Data inválida: {date_str}. Use o formato YYYY-MM-DD")

def build_filial_filter(company: str, filial: Optional[str], cursor) -> str:
    """Constrói filtro SQL para filiais (mesma lógica do TypeScript)"""
    if not company or company not in COMPANIES:
        return ''
    
    company_config = COMPANIES[company]
    filiais = company_config['filiais']
    
    if filial and filial != '__VAREJO__':
        # Filtrar por filial específica
        cursor.execute("SELECT @filial = ?", (filial,))
        return "AND LTRIM(RTRIM(CAST(cv.FILIAL AS VARCHAR))) = LTRIM(RTRIM(CAST(@filial AS VARCHAR)))"
    
    if company == 'scarfme' and filial == '__VAREJO__':
        # Apenas filiais normais (sem ecommerce)
        normal_filiais = [f for f in filiais if f not in company_config.get('ecommerce_filiais', [])]
        if not normal_filiais:
            return ''
        placeholders = ', '.join([f"'{f}'" for f in normal_filiais])
        return f"AND cv.FILIAL IN ({placeholders})"
    
    # Todas as filiais da empresa
    if not filiais:
        return ''
    placeholders = ', '.join([f"'{f}'" for f in filiais])
    return f"AND cv.FILIAL IN ({placeholders})"

def fetch_clientes(
    company: Optional[str] = None,
    filial: Optional[str] = None,
    vendedor: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    search_term: Optional[str] = None
) -> pd.DataFrame:
    """Busca clientes cadastrados no período (mesma lógica do TypeScript)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Usar 2025 inteiro como padrão se não especificado
    if not start_date or not end_date:
        start_date = date(2025, 1, 1)
        end_date = date(2026, 1, 1)  # 2026-01-01 exclusivo = até 2025-12-31
    
    # Adicionar 1 dia ao end_date para incluir todo o dia final (exclusivo)
    end_date_plus_one = end_date + timedelta(days=1)
    
    # Construir filtros
    filial_filter = build_filial_filter(company, filial, cursor) if company else ''
    
    vendedor_filter = ''
    if vendedor and vendedor.strip():
        vendedor_filter = f"AND (LTRIM(RTRIM(CAST(cv.VENDEDOR AS VARCHAR))) = '{vendedor.strip()}' OR LTRIM(RTRIM(ISNULL(lv.VENDEDOR_APELIDO, ISNULL(lv.NOME_VENDEDOR, '')))) = '{vendedor.strip()}')"
    
    search_filter = ''
    if search_term and len(search_term.strip()) >= 2:
        search_pattern = f"%{search_term.strip()}%"
        search_filter = f"AND (cv.CLIENTE_VAREJO LIKE '{search_pattern}' OR ISNULL(lv.VENDEDOR_APELIDO, ISNULL(lv.NOME_VENDEDOR, cv.VENDEDOR)) LIKE '{search_pattern}')"
    
    # Query SQL (mesma do TypeScript) + email
    query = f"""
        SELECT 
            CAST(cv.CADASTRAMENTO AS DATE) AS data,
            ISNULL(cv.CLIENTE_VAREJO, 'SEM NOME') AS nomeCliente,
            CASE 
                WHEN cv.DDD IS NOT NULL AND cv.TELEFONE IS NOT NULL 
                THEN cv.DDD + ' ' + cv.TELEFONE 
                ELSE ISNULL(cv.TELEFONE, '') 
            END AS telefone,
            ISNULL(cv.CPF_CGC, '') AS cpf,
            ISNULL(cv.ENDERECO, '') AS endereco,
            ISNULL(cv.COMPLEMENTO, '') AS complemento,
            ISNULL(cv.BAIRRO, '') AS bairro,
            ISNULL(cv.CIDADE, '') AS cidade,
            ISNULL(cv.EMAIL, '') AS email,
            ISNULL(lv.VENDEDOR_APELIDO, ISNULL(lv.NOME_VENDEDOR, cv.VENDEDOR)) AS vendedor,
            cv.FILIAL AS filial
        FROM CLIENTES_VAREJO cv WITH (NOLOCK)
        LEFT JOIN LOJA_VENDEDORES lv WITH (NOLOCK)
            ON LTRIM(RTRIM(CAST(cv.VENDEDOR AS VARCHAR))) = LTRIM(RTRIM(CAST(lv.VENDEDOR AS VARCHAR)))
        WHERE CAST(cv.CADASTRAMENTO AS DATE) >= '{start_date}'
            AND CAST(cv.CADASTRAMENTO AS DATE) < '{end_date_plus_one}'
            {filial_filter}
            {vendedor_filter}
            {search_filter}
        ORDER BY cv.CADASTRAMENTO ASC, cv.CLIENTE_VAREJO
    """
    
    try:
        df = pd.read_sql(query, conn)
        print(f"✓ {len(df)} clientes encontrados")
        return df
    except Exception as e:
        print(f"Erro ao buscar clientes: {e}")
        raise
    finally:
        conn.close()

def fetch_vendas_clientes(
    company: Optional[str] = None,
    filial: Optional[str] = None,
    vendedor: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    clientes_df: Optional[pd.DataFrame] = None
) -> pd.DataFrame:
    """Busca vendas relacionadas aos clientes cadastrados"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if not start_date or not end_date:
        start_date = date(2025, 1, 1)
        end_date = date(2026, 1, 1)  # 2026-01-01 exclusivo = até 2025-12-31
    
    end_date_plus_one = end_date + timedelta(days=1)
    
    # Construir filtros
    filial_filter = build_filial_filter(company, filial, cursor) if company else ''
    
    vendedor_filter = ''
    if vendedor and vendedor.strip():
        vendedor_filter = f"AND (LTRIM(RTRIM(CAST(vp.VENDEDOR AS VARCHAR))) = '{vendedor.strip()}' OR LTRIM(RTRIM(ISNULL(v.VENDEDOR_APELIDO, '')))) = '{vendedor.strip()}')"
    
    # Se temos lista de clientes, filtrar por eles
    cliente_filter = ''
    if clientes_df is not None and len(clientes_df) > 0:
        clientes_nomes = clientes_df['nomeCliente'].unique()
        if len(clientes_nomes) > 0:
            # Limitar a 1000 clientes para não sobrecarregar a query
            if len(clientes_nomes) > 1000:
                clientes_nomes = clientes_nomes[:1000]
            clientes_list = "', '".join([str(n).replace("'", "''") for n in clientes_nomes])
            cliente_filter = f"AND v.CLIENTE_VAREJO IN ('{clientes_list}')"
    
    # Query para buscar vendas dos clientes
    query = f"""
        SELECT 
            CAST(vp.DATA_VENDA AS DATE) AS dataVenda,
            ISNULL(v.CLIENTE_VAREJO, 'SEM CLIENTE') AS nomeCliente,
            vp.FILIAL AS filial,
            ISNULL(v.VENDEDOR_APELIDO, ISNULL(vp.VENDEDOR, 'SEM VENDEDOR')) AS vendedor,
            v.TICKET AS ticket,
            vp.PRODUTO AS produto,
            ISNULL(vp.DESC_PRODUTO, '') AS descricaoProduto,
            ISNULL(vp.GRUPO_PRODUTO, '') AS grupo,
            ISNULL(vp.SUBGRUPO_PRODUTO, '') AS subgrupo,
            CASE 
                WHEN vp.QTDE_CANCELADA > 0 THEN 0
                ELSE vp.QTDE
            END AS quantidade,
            CASE 
                WHEN vp.QTDE_CANCELADA > 0 THEN 0
                ELSE (vp.PRECO_LIQUIDO * vp.QTDE) - ISNULL(vp.DESCONTO_VENDA, 0)
            END AS valorLiquido,
            v.VALOR_TIKET AS valorTicket
        FROM W_CTB_LOJA_VENDA_PEDIDO_PRODUTO vp WITH (NOLOCK)
        LEFT JOIN W_CTB_LOJA_VENDA_PEDIDO v WITH (NOLOCK)
            ON v.FILIAL = vp.FILIAL 
            AND v.PEDIDO = vp.PEDIDO 
            AND v.TICKET = vp.TICKET
        WHERE vp.DATA_VENDA >= '{start_date}'
            AND vp.DATA_VENDA < '{end_date_plus_one}'
            AND vp.QTDE > 0
            {filial_filter}
            {vendedor_filter}
            {cliente_filter}
        ORDER BY v.CLIENTE_VAREJO, vp.DATA_VENDA DESC, v.TICKET
    """
    
    try:
        df = pd.read_sql(query, conn)
        print(f"✓ {len(df)} vendas encontradas")
        return df
    except Exception as e:
        print(f"Erro ao buscar vendas: {e}")
        raise
    finally:
        conn.close()

def format_excel(workbook, worksheet, df: pd.DataFrame, table_name: str):
    """Formata a planilha Excel com cores, bordas e tabela dinâmica"""
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formatar cabeçalho
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Formatar dados
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border
            
            # Alinhamento baseado no tipo de dado
            if isinstance(cell_value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif isinstance(cell_value, date):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar largura das colunas
    for col_num, column_title in enumerate(df.columns, 1):
        max_length = max(
            len(str(column_title)),
            df[column_title].astype(str).str.len().max() if len(df) > 0 else 0
        )
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width
    
    # Congelar primeira linha
    worksheet.freeze_panes = 'A2'
    
    # Adicionar tabela dinâmica (Excel Table)
    if len(df) > 0:
        table = Table(displayName=table_name, ref=worksheet.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)

def format_currency(value):
    """Formata valor como moeda brasileira"""
    if pd.isna(value) or value == 0:
        return 0.0
    return float(value)

def create_excel_file(
    clientes_df: pd.DataFrame,
    vendas_df: pd.DataFrame,
    output_file: str,
    company: Optional[str] = None
):
    """Cria arquivo Excel com dados de clientes e vendas"""
    print(f"\nGerando arquivo Excel: {output_file}")
    
    # Preparar dados de clientes
    df_clientes = clientes_df.copy()
    
    # Formatar filiais se tiver company config
    if company and company in COMPANIES:
        company_config = COMPANIES[company]
        display_names = company_config.get('filial_display_names', {})
        df_clientes['filialDisplay'] = df_clientes['filial'].map(
            lambda x: display_names.get(x, x)
        )
    else:
        df_clientes['filialDisplay'] = df_clientes['filial']
    
    # Reordenar colunas de clientes
    col_order = ['data', 'nomeCliente', 'telefone', 'email', 'cpf', 'endereco', 
                 'complemento', 'bairro', 'cidade', 'vendedor', 'filial', 'filialDisplay']
    col_order = [c for c in col_order if c in df_clientes.columns]
    df_clientes = df_clientes[col_order]
    
    # Renomear colunas para português
    col_names_pt = []
    col_mapping = {
        'data': 'Data Cadastro',
        'nomeCliente': 'Nome Cliente',
        'telefone': 'Telefone',
        'email': 'E-mail',
        'cpf': 'CPF/CNPJ',
        'endereco': 'Endereço',
        'complemento': 'Complemento',
        'bairro': 'Bairro',
        'cidade': 'Cidade',
        'vendedor': 'Vendedor',
        'filial': 'Filial',
        'filialDisplay': 'Filial (Display)'
    }
    for col in col_order:
        col_names_pt.append(col_mapping.get(col, col))
    df_clientes.columns = col_names_pt
    
    # Preparar dados de vendas
    df_vendas = vendas_df.copy()
    
    # Agrupar vendas por cliente para criar resumo
    if len(df_vendas) > 0:
        vendas_resumo = df_vendas.groupby('nomeCliente').agg({
            'dataVenda': ['min', 'max', 'count'],
            'ticket': 'nunique',
            'valorLiquido': 'sum',
            'quantidade': 'sum'
        }).reset_index()
        
        vendas_resumo.columns = ['Nome Cliente', 'Primeira Venda', 'Última Venda', 
                                 'Total Itens', 'Total Tickets', 'Faturamento Total', 'Quantidade Total']
        
        # Formatar valores monetários
        vendas_resumo['Faturamento Total'] = vendas_resumo['Faturamento Total'].apply(format_currency)
        
        # Detalhes de vendas
        df_vendas_detalhes = df_vendas.copy()
        df_vendas_detalhes = df_vendas_detalhes[[
            'dataVenda', 'nomeCliente', 'filial', 'vendedor', 'ticket',
            'produto', 'descricaoProduto', 'grupo', 'subgrupo', 
            'quantidade', 'valorLiquido'
        ]]
        df_vendas_detalhes.columns = [
            'Data Venda', 'Nome Cliente', 'Filial', 'Vendedor', 'Ticket',
            'Produto', 'Descrição', 'Grupo', 'Subgrupo', 
            'Quantidade', 'Valor Líquido'
        ]
        df_vendas_detalhes['Valor Líquido'] = df_vendas_detalhes['Valor Líquido'].apply(format_currency)
    else:
        vendas_resumo = pd.DataFrame(columns=[
            'Nome Cliente', 'Primeira Venda', 'Última Venda', 
            'Total Itens', 'Total Tickets', 'Faturamento Total', 'Quantidade Total'
        ])
        df_vendas_detalhes = pd.DataFrame()
    
    # Criar arquivo Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Aba 1: Clientes
        df_clientes.to_excel(writer, sheet_name='Clientes', index=False)
        ws_clientes = writer.sheets['Clientes']
        format_excel(writer.book, ws_clientes, df_clientes, 'TabelaClientes')
        
        # Aba 2: Resumo de Vendas por Cliente
        if len(vendas_resumo) > 0:
            vendas_resumo.to_excel(writer, sheet_name='Resumo Vendas', index=False)
            ws_resumo = writer.sheets['Resumo Vendas']
            format_excel(writer.book, ws_resumo, vendas_resumo, 'TabelaResumoVendas')
            
            # Formatar coluna de faturamento como moeda
            for row in range(2, len(vendas_resumo) + 2):
                cell = ws_resumo.cell(row=row, column=6)  # Coluna F (Faturamento Total)
                cell.number_format = 'R$ #,##0.00'
        
        # Aba 3: Detalhes de Vendas
        if len(df_vendas_detalhes) > 0:
            df_vendas_detalhes.to_excel(writer, sheet_name='Detalhes Vendas', index=False)
            ws_detalhes = writer.sheets['Detalhes Vendas']
            format_excel(writer.book, ws_detalhes, df_vendas_detalhes, 'TabelaDetalhesVendas')
            
            # Formatar coluna de valor como moeda
            for row in range(2, len(df_vendas_detalhes) + 2):
                cell = ws_detalhes.cell(row=row, column=11)  # Coluna K (Valor Líquido)
                cell.number_format = 'R$ #,##0.00'
    
    print(f"✓ Arquivo Excel criado com sucesso: {output_file}")
    print(f"  - {len(df_clientes)} clientes")
    print(f"  - {len(vendas_resumo)} clientes com vendas")
    print(f"  - {len(df_vendas_detalhes)} itens de venda")

def main():
    parser = argparse.ArgumentParser(
        description='Exporta dados de clientes e vendas para Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  # Exportar clientes do mês atual (NERD)
  python exportar_clientes.py --company nerd

  # Exportar com filtros específicos
  python exportar_clientes.py --company scarfme --filial "SCARF ME - MATRIZ" --start 2024-01-01 --end 2024-01-31

  # Exportar com busca por termo
  python exportar_clientes.py --company nerd --search "João"
        """
    )
    
    parser.add_argument('--company', choices=['nerd', 'scarfme'], 
                       help='Empresa (nerd ou scarfme)')
    parser.add_argument('--filial', type=str, 
                       help='Filial específica (ou "__VAREJO__" para apenas varejo)')
    parser.add_argument('--vendedor', type=str, 
                       help='Código ou nome do vendedor')
    parser.add_argument('--start', type=str, 
                       help='Data inicial (YYYY-MM-DD). Padrão: primeiro dia do mês atual')
    parser.add_argument('--end', type=str, 
                       help='Data final (YYYY-MM-DD). Padrão: último dia do mês atual')
    parser.add_argument('--search', type=str, 
                       help='Termo de busca (nome do cliente ou vendedor)')
    parser.add_argument('--output', type=str, default=None,
                       help='Nome do arquivo de saída. Padrão: clientes_YYYYMMDD_HHMMSS.xlsx')
    parser.add_argument('--no-vendas', action='store_true',
                       help='Não buscar dados de vendas (apenas clientes)')
    
    args = parser.parse_args()
    
    # Validar datas
    start_date = None
    end_date = None
    if args.start:
        start_date = parse_date(args.start)
    if args.end:
        end_date = parse_date(args.end)
    
    if start_date and end_date and start_date > end_date:
        print("Erro: Data inicial não pode ser maior que data final")
        sys.exit(1)
    
    # Nome do arquivo de saída
    if not args.output:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        company_suffix = f"_{args.company}" if args.company else ""
        args.output = f"clientes{company_suffix}_{timestamp}.xlsx"
    
    try:
        print("=" * 60)
        print("EXPORTAÇÃO DE CLIENTES E VENDAS")
        print("=" * 60)
        print(f"Empresa: {args.company or 'Todas'}")
        print(f"Filial: {args.filial or 'Todas'}")
        print(f"Vendedor: {args.vendedor or 'Todos'}")
        if start_date and end_date:
            print(f"Período: {start_date} até {end_date}")
        else:
            print(f"Período: 2025 inteiro (2025-01-01 até 2025-12-31)")
        print(f"Busca: {args.search or 'Nenhuma'}")
        print("-" * 60)
        
        # Buscar clientes
        print("\n[1/2] Buscando clientes...")
        df_clientes = fetch_clientes(
            company=args.company,
            filial=args.filial,
            vendedor=args.vendedor,
            start_date=start_date,
            end_date=end_date,
            search_term=args.search
        )
        
        if len(df_clientes) == 0:
            print("Nenhum cliente encontrado. Abortando.")
            sys.exit(0)
        
        # Buscar vendas
        df_vendas = pd.DataFrame()
        if not args.no_vendas:
            print("\n[2/2] Buscando vendas...")
            try:
                df_vendas = fetch_vendas_clientes(
                    company=args.company,
                    filial=args.filial,
                    vendedor=args.vendedor,
                    start_date=start_date,
                    end_date=end_date,
                    clientes_df=df_clientes
                )
            except Exception as e:
                print(f"⚠ Aviso: Erro ao buscar vendas: {e}")
                print("  Continuando apenas com dados de clientes...")
                df_vendas = pd.DataFrame()
        
        # Gerar Excel
        print("\n[3/3] Gerando arquivo Excel...")
        create_excel_file(df_clientes, df_vendas, args.output, args.company)
        
        print("\n" + "=" * 60)
        print("✓ EXPORTAÇÃO CONCLUÍDA COM SUCESSO!")
        print("=" * 60)
        
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()

