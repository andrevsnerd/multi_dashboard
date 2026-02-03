#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Relatório de entradas NERD por mês.
Mostra todos os produtos que entraram nas filiais NERD no mês escolhido.

Fontes de entradas (duas no sistema):
  1) ESTOQUE_PROD_ENT + ESTOQUE_PROD1_ENT  (romaneios de entrada “oficiais”)
  2) LOJA_ENTRADAS + LOJA_ENTRADAS_PRODUTO (entradas de loja, ex.: transferências)
O exportador só gera data/entradas.csv da fonte 1. Por isso, se rodar com conexão
ao banco, o script usa as DUAS fontes (sem duplicar romaneios que existem em ambas).
Sem banco, usa apenas data/entradas.csv e pode faltar itens.
"""

import pandas as pd
import os
import re
from datetime import datetime

# Caminhos relativos ao script (pasta multi-dashboard)
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
RELATORIOS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'relatorios')

# Config banco (mesma do exportador) – usado para buscar as duas fontes de entradas
DB_CONFIG = {
    'server': '177.92.78.250',
    'server_fallback': '189.126.197.82',
    'database': 'LINX_PRODUCAO',
    'username': 'andre.nerd',
    'password': 'nerd123@'
}


def _parse_decimal(series):
    """Converte coluna com vírgula ou ponto decimal para float."""
    return pd.to_numeric(series.astype(str).str.replace(',', '.', regex=False), errors='coerce').fillna(0)


def conectar_banco():
    """Conecta ao SQL Server (mesma config do exportador). Retorna None se falhar."""
    try:
        import pyodbc
    except ImportError:
        return None
    for nome, servidor in [('principal', DB_CONFIG['server']), ('fallback', DB_CONFIG['server_fallback'])]:
        try:
            conn_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={servidor};"
                f"DATABASE={DB_CONFIG['database']};"
                f"UID={DB_CONFIG['username']};"
                f"PWD={DB_CONFIG['password']};"
                f"Connection Timeout=30;"
            )
            conn = pyodbc.connect(conn_str)
            conn.timeout = 300
            return conn
        except Exception:
            continue
    return None


def fetch_entradas_banco(conn, ano, mes):
    """
    Busca entradas NERD no mês de DUAS fontes (sem duplicar):
    1) ESTOQUE_PROD_ENT + ESTOQUE_PROD1_ENT
    2) LOJA_ENTRADAS + LOJA_ENTRADAS_PRODUTO (apenas romaneios que NÃO existem em ESTOQUE_PROD_ENT)
    Retorna DataFrame com: EMISSAO, FILIAL, PRODUTO, COR_PRODUTO, QTDE_TOTAL, DESC_PRODUTO, DESC_COR_PRODUTO
    """
    from datetime import date
    start = date(ano, mes, 1)
    if mes == 12:
        end = date(ano + 1, 1, 1)
    else:
        end = date(ano, mes + 1, 1)
    start_s = start.isoformat()
    end_s = end.isoformat()

    # Fonte 1: ESTOQUE_PROD_ENT
    q1 = """
    SELECT E.EMISSAO, E.FILIAL, P.PRODUTO, P.COR_PRODUTO, CAST(P.QTDE AS FLOAT) AS QTDE_TOTAL,
           pr.DESC_PRODUTO, c.DESC_COR AS DESC_COR_PRODUTO
    FROM ESTOQUE_PROD_ENT AS E WITH (NOLOCK)
    LEFT JOIN ESTOQUE_PROD1_ENT AS P WITH (NOLOCK) ON E.ROMANEIO_PRODUTO = P.ROMANEIO_PRODUTO
    LEFT JOIN PRODUTOS pr WITH (NOLOCK) ON pr.PRODUTO = P.PRODUTO
    LEFT JOIN CORES_BASICAS c WITH (NOLOCK) ON c.COR = P.COR_PRODUTO
    WHERE E.EMISSAO >= ? AND E.EMISSAO < ?
      AND E.FILIAL LIKE '%NERD%'
      AND P.PRODUTO IS NOT NULL
    """
    # Fonte 2: LOJA_ENTRADAS (só romaneios que não estão em ESTOQUE_PROD_ENT)
    q2 = """
    SELECT LE.EMISSAO, LE.FILIAL, LEP.PRODUTO, LEP.COR_PRODUTO,
           CAST(ISNULL(LEP.QTDE_ENTRADA, 0) AS FLOAT) AS QTDE_TOTAL,
           pr.DESC_PRODUTO, c.DESC_COR AS DESC_COR_PRODUTO
    FROM LOJA_ENTRADAS AS LE WITH (NOLOCK)
    INNER JOIN LOJA_ENTRADAS_PRODUTO AS LEP WITH (NOLOCK)
      ON LEP.FILIAL = LE.FILIAL AND LEP.ROMANEIO_PRODUTO = LE.ROMANEIO_PRODUTO
    LEFT JOIN PRODUTOS pr WITH (NOLOCK) ON pr.PRODUTO = LEP.PRODUTO
    LEFT JOIN CORES_BASICAS c WITH (NOLOCK) ON c.COR = LEP.COR_PRODUTO
    WHERE LE.EMISSAO >= ? AND LE.EMISSAO < ?
      AND LE.FILIAL LIKE '%NERD%'
      AND LEP.PRODUTO IS NOT NULL
      AND NOT EXISTS (
        SELECT 1 FROM ESTOQUE_PROD_ENT E WITH (NOLOCK)
        WHERE E.FILIAL = LE.FILIAL AND E.ROMANEIO_PRODUTO = LE.ROMANEIO_PRODUTO
          AND E.EMISSAO >= ? AND E.EMISSAO < ?
      )
    """
    try:
        df1 = pd.read_sql(q1, conn, params=(start_s, end_s))
        df2 = pd.read_sql(q2, conn, params=(start_s, end_s, start_s, end_s))
    except Exception as e:
        return None, str(e)
    df1['_fonte'] = 'ESTOQUE_PROD_ENT'
    df2['_fonte'] = 'LOJA_ENTRADAS'
    # Normalizar COR_PRODUTO para string
    for df in (df1, df2):
        if 'COR_PRODUTO' in df.columns:
            df['COR_PRODUTO'] = df['COR_PRODUTO'].fillna('').astype(str).str.strip()
    out = pd.concat([df1, df2], ignore_index=True)
    return out, None


def _slug_filename(texto: str) -> str:
    """Gera um 'slug' seguro para nome de arquivo (Windows)."""
    if texto is None:
        return ''
    s = str(texto).strip()
    s = re.sub(r'[<>:"/\\|?*]+', '', s)
    s = re.sub(r'\s+', '_', s)
    s = re.sub(r'_+', '_', s)
    return s.strip('._-')


def obter_mes_ano():
    """
    Pede ao usuário o mês no formato MM-AA (ex: 01-26 para janeiro/2026).
    Retorna (ano: int, mes: int).
    """
    print("\n" + "="*60)
    print("ENTRADAS NERD - POR MÊS")
    print("="*60)
    print("Informe o mês no formato MM-AA (ex: 01-26 = janeiro/2026)")
    print("-"*60)
    while True:
        entrada = input("Mês [MM-AA]: ").strip()
        if not entrada:
            # Padrão: mês atual
            hoje = datetime.now()
            return hoje.year, hoje.month
        match = re.match(r'^(\d{1,2})-(\d{2,4})$', entrada)
        if match:
            mes = int(match.group(1))
            aa = match.group(2)
            ano = 2000 + int(aa) if len(aa) == 2 else int(aa)
            if 1 <= mes <= 12 and 2000 <= ano <= 2100:
                return ano, mes
        print("⚠️ Formato inválido. Use MM-AA (ex: 01-26).")


def aplicar_formatacao_excel(ws, relatorio, titulo):
    """Aplica formatação de título, cabeçalho e colunas monetárias (estilo estoque_produto_geral_nerd)."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    ws.insert_rows(1)
    max_col_letter = get_column_letter(ws.max_column)
    ws.merge_cells(f'A1:{max_col_letter}1')
    ws['A1'] = titulo
    ws['A1'].fill = PatternFill(start_color='262626', end_color='262626', fill_type='solid')
    ws['A1'].font = Font(color='FFFFFF', size=26, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    titulo_fill = PatternFill(start_color='262626', end_color='262626', fill_type='solid')
    titulo_font = Font(color='FFFFFF', bold=True)
    for col in ws.iter_cols(min_row=2, max_row=2):
        for cell in col:
            cell.fill = titulo_fill
            cell.font = titulo_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 38

    # Largura das colunas e formato R$ nas colunas de valor (row 1 está merged, cabeçalho na row 2)
    colunas_valor = ('Custo Unitário', 'Preço de Venda', 'Custo Total', 'Valor Venda Total')
    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        if col:
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = min(
                max(len(str(c.value or '')) for c in col) + 2, 50
            )
            # Cabeçalho está na linha 2 (col[1] pois col[0] é linha 1 merged)
            header_val = col[1].value if len(col) > 1 else None
            if header_val in colunas_valor:
                for cell in list(col)[2:]:  # a partir da linha 3 (dados)
                    cell.number_format = 'R$ #,##0.00'
            if header_val == 'Markup %':
                for cell in list(col)[2:]:
                    cell.number_format = '0.0%'


def gerar_relatorio():
    """Carrega entradas e produtos, filtra por NERD e mês, gera planilha com produto, descrição, cor, qtde, custo, preço, markup."""
    ano, mes = obter_mes_ano()
    mes_nome_pt = datetime(ano, mes, 1).strftime('%m/%Y')  # ex: 01/2026

    print("\nCarregando dados...")
    produtos_path = os.path.join(DATA_DIR, 'produtos_tratados.csv')
    if not os.path.exists(produtos_path):
        print(f"✗ Arquivo não encontrado: {produtos_path}")
        return

    try:
        produtos = pd.read_csv(produtos_path, sep=';', encoding='utf-8',
                               dtype={'PRODUTO': str}, low_memory=False)
    except Exception as e:
        print(f"✗ Erro ao ler produtos: {e}")
        return
    for c in ['PRODUTO', 'CUSTO_REPOSICAO1', 'PRECO_REPOSICAO_1']:
        if c not in produtos.columns:
            print(f"✗ Coluna faltando em produtos: {c}")
            return

    # Entradas: tentar banco (duas fontes) primeiro; senão CSV (só ESTOQUE_PROD_ENT)
    entradas_nerd = None
    fonte_info = " (fonte: banco ou CSV)"
    conn = conectar_banco()
    if conn:
        print("  Conectado ao banco. Buscando entradas (ESTOQUE_PROD_ENT + LOJA_ENTRADAS)...")
        entradas_nerd, err = fetch_entradas_banco(conn, ano, mes)
        try:
            conn.close()
        except Exception:
            pass
        if err:
            print(f"  ⚠️ Erro na query: {err}. Usando CSV.")
            entradas_nerd = None
        elif entradas_nerd is not None and not entradas_nerd.empty:
            n_ep = (entradas_nerd['_fonte'] == 'ESTOQUE_PROD_ENT').sum()
            n_le = (entradas_nerd['_fonte'] == 'LOJA_ENTRADAS').sum()
            fonte_info = f" (banco: {n_ep:,} ESTOQUE_PROD_ENT + {n_le:,} LOJA_ENTRADAS)"
            entradas_nerd = entradas_nerd.drop(columns=['_fonte'], errors='ignore')
            entradas_nerd['EMISSAO'] = pd.to_datetime(entradas_nerd['EMISSAO'], errors='coerce')
            entradas_nerd['FILIAL'] = entradas_nerd['FILIAL'].astype(str).str.strip()
            entradas_nerd['PRODUTO'] = entradas_nerd['PRODUTO'].astype(str).str.strip()
            if 'DESC_COR_PRODUTO' not in entradas_nerd.columns:
                entradas_nerd['DESC_COR_PRODUTO'] = ''
            if 'DESC_PRODUTO' not in entradas_nerd.columns:
                entradas_nerd['DESC_PRODUTO'] = ''
    if entradas_nerd is None or entradas_nerd.empty:
        # Fallback: CSV (apenas ESTOQUE_PROD_ENT)
        entradas_path = os.path.join(DATA_DIR, 'entradas.csv')
        if not os.path.exists(entradas_path):
            print(f"✗ Arquivo não encontrado: {entradas_path}")
            print("  Execute o exportador para gerar data/entradas.csv ou use conexão ao banco.")
            return
        try:
            entradas = pd.read_csv(entradas_path, sep=';', encoding='utf-8',
                                   dtype={'PRODUTO': str, 'COR_PRODUTO': str}, low_memory=False)
        except Exception as e:
            print(f"✗ Erro ao ler entradas: {e}")
            return
        cols_ent = ['EMISSAO', 'FILIAL', 'PRODUTO', 'DESC_PRODUTO', 'COR_PRODUTO', 'DESC_COR_PRODUTO', 'QTDE_TOTAL']
        if any(c not in entradas.columns for c in cols_ent):
            print(f"✗ Colunas faltando em entradas. Use banco ou exportador atualizado.")
            return
        entradas = entradas.copy()
        entradas['FILIAL'] = entradas['FILIAL'].astype(str).str.replace('\xa0', ' ', regex=False).str.strip()
        mask_nerd = entradas['FILIAL'].str.upper().str.contains('NERD', na=False)
        entradas_nerd = entradas[mask_nerd].copy()
        if entradas_nerd.empty:
            print("✗ Nenhuma entrada encontrada para filiais NERD.")
            return
        entradas_nerd['EMISSAO'] = pd.to_datetime(entradas_nerd['EMISSAO'], errors='coerce')
        entradas_nerd = entradas_nerd.dropna(subset=['EMISSAO'])
        entradas_nerd = entradas_nerd[
            (entradas_nerd['EMISSAO'].dt.year == ano) &
            (entradas_nerd['EMISSAO'].dt.month == mes)
        ]
        fonte_info = " (apenas CSV: ESTOQUE_PROD_ENT – pode faltar itens de LOJA_ENTRADAS)"
    if entradas_nerd.empty:
        print(f"✗ Nenhuma entrada NERD no mês {mes:02d}/{ano}.")
        return
    print(f"✓ Entradas carregadas{fonte_info}")

    # Normalizar produto e cor para merge
    entradas_nerd['PRODUTO'] = entradas_nerd['PRODUTO'].astype(str).str.strip()
    entradas_nerd['COR_PRODUTO'] = entradas_nerd['COR_PRODUTO'].fillna('').astype(str).str.strip()
    produtos['PRODUTO'] = produtos['PRODUTO'].astype(str).str.strip()

    # Custo e preço (vírgula decimal)
    produtos['CUSTO_REPOSICAO1'] = _parse_decimal(produtos['CUSTO_REPOSICAO1'])
    produtos['PRECO_REPOSICAO_1'] = _parse_decimal(produtos['PRECO_REPOSICAO_1'])

    # Merge com produtos para obter custo e preço
    prod_cols = ['PRODUTO', 'CUSTO_REPOSICAO1', 'PRECO_REPOSICAO_1']
    entradas_nerd = entradas_nerd.merge(
        produtos[prod_cols].drop_duplicates(subset=['PRODUTO']),
        on='PRODUTO',
        how='left'
    )

    # Quantidade numérica
    entradas_nerd['QTDE_TOTAL'] = pd.to_numeric(entradas_nerd['QTDE_TOTAL'], errors='coerce').fillna(0)

    # Agrupar por produto + descrição + cor (soma de quantidades; custo/preço do cadastro)
    agg = {
        'QTDE_TOTAL': 'sum',
        'DESC_PRODUTO': 'first',
        'DESC_COR_PRODUTO': 'first',
        'CUSTO_REPOSICAO1': 'first',
        'PRECO_REPOSICAO_1': 'first',
    }
    relatorio = (
        entradas_nerd
        .groupby(['PRODUTO', 'COR_PRODUTO'], as_index=False, observed=True)
        .agg(agg)
        .rename(columns={'QTDE_TOTAL': 'Quantidade'})
    )

    # Nomes amigáveis para cor (usar descrição da cor quando existir)
    relatorio['Cor'] = relatorio['DESC_COR_PRODUTO'].fillna(relatorio['COR_PRODUTO']).str.strip()
    relatorio['Descrição'] = relatorio['DESC_PRODUTO'].fillna('').astype(str).str.strip()
    relatorio['Produto'] = relatorio['PRODUTO']

    # Custo unitário e preço de venda (do cadastro)
    relatorio['Custo Unitário'] = relatorio['CUSTO_REPOSICAO1']
    relatorio['Preço de Venda'] = relatorio['PRECO_REPOSICAO_1']

    # Markup simples: (preço / custo) - 1 em percentual; se custo 0, deixar vazio ou 0
    custo = relatorio['Custo Unitário'].replace(0, float('nan'))
    relatorio['Markup %'] = (relatorio['Preço de Venda'] / custo - 1).fillna(0)

    # Totais em valor
    relatorio['Custo Total'] = (relatorio['Quantidade'] * relatorio['Custo Unitário']).round(2)
    relatorio['Valor Venda Total'] = (relatorio['Quantidade'] * relatorio['Preço de Venda']).round(2)

    # Ordem das colunas de saída
    cols_saida = [
        'Produto', 'Descrição', 'Cor', 'Quantidade',
        'Custo Unitário', 'Preço de Venda', 'Markup %',
        'Custo Total', 'Valor Venda Total'
    ]
    relatorio = relatorio[[c for c in cols_saida if c in relatorio.columns]]
    relatorio = relatorio.sort_values(['Produto', 'Cor'])

    # Totais gerais
    total_custo = relatorio['Custo Total'].sum()
    total_qtde = relatorio['Quantidade'].sum()

    # Salvar Excel
    os.makedirs(RELATORIOS_DIR, exist_ok=True)
    nome_arquivo = f"entradas_nerd_{ano}{mes:02d}.xlsx"
    caminho = os.path.join(RELATORIOS_DIR, nome_arquivo)

    titulo = f"Entradas NERD - {mes_nome_pt}"
    try:
        print(f"\nSalvando: {caminho}")
        with pd.ExcelWriter(caminho, engine='openpyxl', mode='w') as writer:
            relatorio.to_excel(writer, sheet_name='Entradas', index=False)
            ws = writer.sheets['Entradas']
            aplicar_formatacao_excel(ws, relatorio, titulo)
        print(f"✓ Salvo: {caminho}")
    except PermissionError:
        base, ext = os.path.splitext(caminho)
        caminho = f"{base}_{datetime.now():%Y%m%d_%H%M%S}{ext}"
        print(f"⚠️ Arquivo aberto, criando: {caminho}")
        with pd.ExcelWriter(caminho, engine='openpyxl', mode='w') as writer:
            relatorio.to_excel(writer, sheet_name='Entradas', index=False)
            aplicar_formatacao_excel(writer.sheets['Entradas'], relatorio, titulo)

    print("\n" + "="*60)
    print("RESUMO")
    print("="*60)
    print(f"  Período: {mes_nome_pt}")
    print(f"  Fonte: {fonte_info.strip() or 'entradas'}")
    print(f"  Linhas (produto+cor): {len(relatorio):,}")
    print(f"  Quantidade total: {total_qtde:,.0f}")
    print(f"  Total gasto (custo): R$ {total_custo:,.2f}")
    print("="*60)
    print("CONCLUÍDO!")
    print("="*60)


if __name__ == "__main__":
    try:
        gerar_relatorio()
    except Exception as e:
        print(f"\n✗ Erro: {e}")
        import traceback
        traceback.print_exc()
