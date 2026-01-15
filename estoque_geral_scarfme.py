#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Relatório estoque Scarfme - Custo Médio Ponderado"""

import pandas as pd
import os
from datetime import datetime
import re

# Constantes
MESES = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
ANO_ATUAL = datetime.now().year
MES_ATUAL = datetime.now().month
COLUNAS_MESES = MESES[:MES_ATUAL]

FILIAIS = {
    'GUARULHOS - RSR': 'GUARULHOS',
    'IGUATEMI SP - JJJ': 'IGUATEMI',
    'MORUMBI - JJJ': 'MORUMBI',
    'OSCAR FREIRE - FSZ': 'OSCAR FREIRE',
    'SCARF ME - HIGIENOPOLIS 2': 'HIGIENÓPOLIS',
    'SCARFME - IBIRAPUERA LLL': 'IBIRAPUERA',
    'SCARFME ME - PAULISTA FFF': 'PAULISTA',
    'SCARF ME - PAULISTA RSR': 'PAULISTA',
    'VILLA LOBOS - LLL': 'VILLA LOBOS',
    'SCARFME MATRIZ CMS': 'E-COMMERCE',
    'SCARF ME - MATRIZ LLL': 'E-COMMERCE'
}

def obter_tipo_analise():
    """
    Define se a análise é geral (padrão) ou filtrada.
    - Geral: comportamento atual (sem filtros)
    - Filtrada: aplica filtros de item (linha/coleção/subgrupo/grade/descrição/produtos)
    """
    print("\n" + "="*60)
    print("TIPO DE ANÁLISE")
    print("="*60)
    print("1) Geral (padrão)")
    print("2) Filtrada")
    print("-"*60)
    escolha = input("Escolha [1/2] (padrão: 1): ").strip().lower()
    if escolha in ("2", "f", "filtrada"):
        return "filtrada"
    return "geral"

def obter_modo_combinacao_filtros():
    """
    Define como múltiplos filtros devem interagir entre si.
    - acumulativo: OR (amplia)  -> entra se atender QUALQUER filtro informado
    - filtrativo : AND (cruza)  -> entra só se atender TODOS os filtros informados
    """
    print("\n" + "="*60)
    print("MODO DOS FILTROS")
    print("="*60)
    print("1) Acumulativo (ampliar)  → aparece se atender QUALQUER filtro")
    print("2) Filtrativo (cruzar)    → aparece só se atender TODOS os filtros")
    print("-"*60)
    while True:
        escolha = input("Escolha [1/2] (padrão: 2): ").strip().lower()
        if escolha in ("", "2", "f", "filtrativo", "cruzar", "and", "intersecao", "interseção"):
            return "filtrativo"
        if escolha in ("1", "a", "acumulativo", "ampliar", "or", "uniao", "união"):
            return "acumulativo"
        print("✗ Opção inválida. Digite 1 (Acumulativo) ou 2 (Filtrativo).")

def obter_filtros_item():
    """
    Solicita filtros de item (mesmo pacote usado nos outros relatórios SCARFME).
    Aceita múltiplos valores separados por vírgula.
    """
    print("\n" + "="*60)
    print("FILTROS (ITEM)")
    print("="*60)
    print("Deixe vazio para NÃO filtrar aquele campo")
    print("Dica: múltiplos valores separados por vírgula (ex: U7, U8)")
    print("-"*60)
    filtros = {}

    colecao = input("Coleção (ex: U7, U8): ").strip()
    if colecao:
        filtros['colecao'] = [c.strip().upper() for c in colecao.split(',') if c.strip()]

    linha = input("Linha: ").strip()
    if linha:
        filtros['linha'] = [l.strip().upper() for l in linha.split(',') if l.strip()]

    subgrupo = input("Subgrupo/Categoria: ").strip()
    if subgrupo:
        filtros['subgrupo'] = [s.strip().upper() for s in subgrupo.split(',') if s.strip()]

    grade = input("Grade: ").strip()
    if grade:
        filtros['grade'] = [g.strip().upper() for g in grade.split(',') if g.strip()]

    descricao = input("Palavras na descrição (ex: lenço, echarpe): ").strip()
    if descricao:
        filtros['descricao'] = [d.strip().upper() for d in descricao.split(',') if d.strip()]

    produtos = input("Códigos de produto (coluna PRODUTO, separados por vírgula): ").strip()
    if produtos:
        filtros['produtos'] = [p.strip() for p in produtos.split(',') if p.strip()]

    if not filtros:
        print("✓ Nenhum filtro informado")
    return filtros

def aplicar_filtros_dinamicos(df, filtros, modo_combinacao: str = 'filtrativo'):
    """
    Aplica filtros (linha/coleção/subgrupo/grade/descrição/produtos) com modo:
      - filtrativo (AND): precisa atender TODOS os filtros informados
      - acumulativo (OR): precisa atender QUALQUER filtro informado
    """
    if df is None:
        return df
    if not filtros:
        return df
    if getattr(df, 'empty', False):
        return df

    modo = (modo_combinacao or 'filtrativo').strip().lower()
    if modo in ('intersecao', 'interseção', 'and'):
        modo = 'filtrativo'
    if modo in ('uniao', 'união', 'or'):
        modo = 'acumulativo'
    if modo not in ('filtrativo', 'acumulativo'):
        modo = 'filtrativo'

    df_filtrado = df.copy()
    masks = []

    def _add_mask(mask):
        if mask is None:
            return
        masks.append(mask.fillna(False))

    # Coleção
    if 'colecao' in filtros and 'COLECAO' in df_filtrado.columns:
        colecoes = [str(c).strip().upper() for c in filtros.get('colecao', []) if str(c).strip()]
        if colecoes:
            df_filtrado['COLECAO'] = df_filtrado['COLECAO'].astype(str).str.strip().str.upper()
            _add_mask(df_filtrado['COLECAO'].isin(colecoes))

    # Linha
    if 'linha' in filtros and 'LINHA' in df_filtrado.columns:
        linhas = [str(l).strip().upper() for l in filtros.get('linha', []) if str(l).strip()]
        if linhas:
            df_filtrado['LINHA'] = df_filtrado['LINHA'].astype(str).str.strip().str.upper()
            _add_mask(df_filtrado['LINHA'].isin(linhas))

    # Subgrupo
    if 'subgrupo' in filtros and 'SUBGRUPO_PRODUTO' in df_filtrado.columns:
        subgrupos = [str(s).strip().upper() for s in filtros.get('subgrupo', []) if str(s).strip()]
        if subgrupos:
            df_filtrado['SUBGRUPO_PRODUTO'] = df_filtrado['SUBGRUPO_PRODUTO'].astype(str).str.strip().str.upper()
            _add_mask(df_filtrado['SUBGRUPO_PRODUTO'].isin(subgrupos))

    # Grade
    if 'grade' in filtros and 'GRADE' in df_filtrado.columns:
        grades = [str(g).strip().upper() for g in filtros.get('grade', []) if str(g).strip()]
        if grades:
            df_filtrado['GRADE'] = df_filtrado['GRADE'].astype(str).str.strip().str.upper()
            _add_mask(df_filtrado['GRADE'].isin(grades))

    # Descrição (qualquer palavra)
    if 'descricao' in filtros and 'DESC_PRODUTO' in df_filtrado.columns:
        palavras = [str(p).strip().upper() for p in filtros.get('descricao', []) if str(p).strip()]
        if palavras:
            df_filtrado['DESC_PRODUTO'] = df_filtrado['DESC_PRODUTO'].astype(str).str.strip().str.upper()
            pattern = '|'.join([re.escape(p) for p in palavras])
            _add_mask(df_filtrado['DESC_PRODUTO'].str.contains(pattern, na=False, regex=True))

    # Produtos
    if 'produtos' in filtros and 'PRODUTO' in df_filtrado.columns:
        codigos = [str(c).strip() for c in filtros.get('produtos', []) if str(c).strip()]
        if codigos:
            produto_str = df_filtrado['PRODUTO'].astype(str).str.strip()
            _add_mask(produto_str.isin(codigos))

    if not masks:
        return df_filtrado

    if modo == 'acumulativo':
        mascara_final = pd.Series(False, index=df_filtrado.index)
        for m in masks:
            mascara_final |= m
    else:
        mascara_final = pd.Series(True, index=df_filtrado.index)
        for m in masks:
            mascara_final &= m

    return df_filtrado[mascara_final].copy()

def _slug_filename(texto: str) -> str:
    """Gera um 'slug' seguro para nome de arquivo (Windows)."""
    if texto is None:
        return ''
    s = str(texto).strip()
    # Remover caracteres proibidos no Windows: <>:"/\|?*
    s = re.sub(r'[<>:"/\\\\|?*]+', '', s)
    # Normalizar espaços e separadores
    s = re.sub(r'\s+', '_', s)
    s = re.sub(r'_+', '_', s)
    return s.strip('._-')

def gerar_nome_arquivo(tipo_analise: str, modo_filtros: str = None, filtros_item: dict = None) -> str:
    """
    Monta nome de arquivo dinâmico baseado no tipo de análise e filtros.
    """
    partes = ["estoque", "scarfme", str(ANO_ATUAL)]
    if (tipo_analise or "").lower() == "filtrada":
        partes.append("filtrada")
        if modo_filtros:
            partes.append(_slug_filename(modo_filtros))

        filtros_item = filtros_item or {}
        ordem = [
            ("linha", "LIN"),
            ("colecao", "COL"),
            ("subgrupo", "SUB"),
            ("grade", "GRD"),
            ("descricao", "DESC"),
            ("produtos", "PROD"),
        ]
        for chave, prefixo in ordem:
            if chave in filtros_item and filtros_item[chave]:
                vals = filtros_item[chave]
                if not isinstance(vals, list):
                    vals = [vals]
                vals = [_slug_filename(v) for v in vals if _slug_filename(v)]
                if not vals:
                    continue
                # Limitar para não estourar nome
                vals = vals[:3]
                partes.append(f"{prefixo}-{ '-'.join(vals) }")
    else:
        partes.append("geral")

    nome_base = "_".join([p for p in partes if p])
    # Limitar tamanho do nome (evitar path longo no Windows)
    if len(nome_base) > 150:
        nome_base = nome_base[:150]
    return os.path.join("relatorios", f"{nome_base}.xlsx")

def processar_vendas_por_filial(vendas, ecommerce, filial_nome):
    """Processa vendas de uma filial específica"""
    vendas_filial = pd.DataFrame()
    
    if filial_nome == 'E-COMMERCE' and ecommerce is not None:
        vendas_filial = (ecommerce
            .assign(
                DATA_VENDA=lambda x: pd.to_datetime(x['DATA_SAIDA'], errors='coerce'),
                LINHA=lambda x: x['LINHA'].str.strip(),
                SUBGRUPO_PRODUTO=lambda x: x['SUBGRUPO_PRODUTO'].str.strip(),
                GRADE=lambda x: x['GRADE'].str.strip(),
                QTDE=lambda x: pd.to_numeric(x['QTDE'], errors='coerce').fillna(0)
            )
            .dropna(subset=['DATA_VENDA'])
            .query(f'DATA_VENDA.dt.year == {ANO_ATUAL} and QTDE > 0')
            .assign(MES=lambda x: x['DATA_VENDA'].dt.month)
        )
    elif vendas is not None:
        # Mapear nome da filial para o formato do banco
        filial_original = next((k for k, v in FILIAIS.items() if v == filial_nome), filial_nome)
        vendas_filial = (vendas
            .assign(
                DATA_VENDA=lambda x: pd.to_datetime(x['DATA_VENDA'], errors='coerce'),
                LINHA=lambda x: x['LINHA'].str.strip(),
                SUBGRUPO_PRODUTO=lambda x: x['SUBGRUPO_PRODUTO'].str.strip(),
                GRADE=lambda x: x['GRADE'].str.strip(),
                QTDE=lambda x: pd.to_numeric(x['QTDE'], errors='coerce').fillna(0)
            )
            .dropna(subset=['DATA_VENDA'])
            .query(f'DATA_VENDA.dt.year == {ANO_ATUAL} and QTDE > 0 and FILIAL == "{filial_original}"')
            .assign(MES=lambda x: x['DATA_VENDA'].dt.month)
        )
    
    return vendas_filial[['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE', 'MES', 'QTDE']] if not vendas_filial.empty else pd.DataFrame()

def calcular_vendas_mensais_filial(vendas, ecommerce, filial_nome):
    """Calcula vendas mensais por categoria para uma filial específica"""
    vendas_filial = processar_vendas_por_filial(vendas, ecommerce, filial_nome)
    
    if vendas_filial.empty:
        return pd.DataFrame()
    
    # Pivot direto
    vendas_pivot = (vendas_filial
        .groupby(['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE', 'MES'], observed=True)['QTDE']
        .sum()
        .unstack(fill_value=0)
        .reset_index()
        .rename(columns={i: COLUNAS_MESES[i-1] for i in range(1, MES_ATUAL+1) if i in range(1, 13)})
    )
    
    # Adiciona meses faltantes
    for col in COLUNAS_MESES:
        if col not in vendas_pivot.columns:
            vendas_pivot[col] = 0
    
    return vendas_pivot

def aplicar_formatacao_cores(ws, relatorio, titulo="Estoque Total"):
    """Aplica formatação condicional e estilos"""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.formatting.rule import CellIsRule
    
    # Título principal mesclado
    from openpyxl.utils import get_column_letter
    ws.insert_rows(1)
    max_col_letter = get_column_letter(ws.max_column)
    ws.merge_cells(f'A1:{max_col_letter}1')
    ws['A1'] = titulo
    ws['A1'].fill = PatternFill(start_color='262626', end_color='262626', fill_type='solid')
    ws['A1'].font = Font(color='FFFFFF', size=26, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Formatação dos títulos das colunas
    titulo_fill = PatternFill(start_color='262626', end_color='262626', fill_type='solid')
    titulo_font = Font(color='FFFFFF', bold=True)
    
    for col in ws.iter_cols(min_row=2, max_row=2):
        for cell in col:
            cell.fill = titulo_fill
            cell.font = titulo_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Altura do cabeçalho
    ws.row_dimensions[2].height = 38
    
    # Formatação condicional para Dias de Estoque
    if 'Dias de Estoque' not in relatorio.columns:
        return
    
    col_idx = list(relatorio.columns).index('Dias de Estoque') + 1
    col_letter = get_column_letter(col_idx)
    range_cells = f'{col_letter}3:{col_letter}{len(relatorio)+2}'
    
    verde = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    vermelho = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    
    ws.conditional_formatting.add(range_cells, CellIsRule(operator='lessThanOrEqual', formula=['90'], fill=verde))
    ws.conditional_formatting.add(range_cells, CellIsRule(operator='between', formula=['90', '180'], fill=amarelo))
    ws.conditional_formatting.add(range_cells, CellIsRule(operator='greaterThan', formula=['180'], fill=vermelho))

def processar_filial(estoque, vendas, ecommerce, filial_nome, filial_original):
    """Processa dados de uma filial específica"""
    print(f"  Processando {filial_nome}...")
    
    # Filtra estoque da filial
    estoque_filial = estoque[estoque['FILIAL'] == filial_original].copy()
    
    if estoque_filial.empty:
        print(f"    ⚠️ Sem estoque para {filial_nome}")
        return None
    
    # Agrupamento
    relatorio = (estoque_filial
        .groupby(['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE'], observed=True)
        .agg({'CUSTO_TOTAL_ITEM': 'sum', 'ESTOQUE': 'sum'})
        .reset_index()
        .assign(CUSTO_UNITARIO=lambda x: x['CUSTO_TOTAL_ITEM'] / x['ESTOQUE'])
    )
    
    # Integra vendas se disponível
    vendas_mensais = calcular_vendas_mensais_filial(vendas, ecommerce, filial_nome)
    if not vendas_mensais.empty:
        relatorio = relatorio.merge(vendas_mensais, on=['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE'], how='left')
        
        # Preenche NaN
        for col in COLUNAS_MESES:
            if col in relatorio.columns:
                relatorio[col] = relatorio[col].fillna(0)
        
        # Métricas
        meses_calc = min(3, len(COLUNAS_MESES))
        ultimos_meses = COLUNAS_MESES[-meses_calc:] if meses_calc > 0 else []
        
        if ultimos_meses:
            relatorio['Projeção Mensal'] = relatorio[ultimos_meses].mean(axis=1).round(0)
            meses_restantes = max(12 - MES_ATUAL + 1, 0)
            relatorio['Projeção Ano'] = (relatorio['Projeção Mensal'] * meses_restantes).round(0)
            relatorio['Estoque Final Mês'] = (relatorio['ESTOQUE'] - relatorio['Projeção Mensal']).round(0)
            relatorio['Estoque Final Ano'] = (relatorio['ESTOQUE'] - relatorio['Projeção Ano']).round(0)
            relatorio['Dias de Estoque'] = (relatorio['ESTOQUE'] / relatorio['Projeção Mensal'].replace(0, 1) * 30).round(1)
            relatorio.loc[relatorio['Projeção Mensal'] == 0, 'Dias de Estoque'] = 999
    
    # Renomeia e ordena
    relatorio = (relatorio
        .rename(columns={
            'LINHA': 'Linha',
            'SUBGRUPO_PRODUTO': 'Subgrupo',
            'GRADE': 'Grade',
            'CUSTO_UNITARIO': 'Custo Unitário',
            'CUSTO_TOTAL_ITEM': 'Custo Total',
            'ESTOQUE': 'Estoque Total'
        })
    )
    
    # Ordem das colunas
    cols_ordem = ['Linha', 'Subgrupo', 'Grade', 'Custo Unitário', 'Custo Total', 'Estoque Total']
    if not vendas_mensais.empty:
        cols_ordem.extend(COLUNAS_MESES + ['Projeção Mensal', 'Projeção Ano', 'Estoque Final Mês', 'Estoque Final Ano', 'Dias de Estoque'])
    
    relatorio = relatorio[[c for c in cols_ordem if c in relatorio.columns]]
    relatorio[['Custo Unitário', 'Custo Total']] = relatorio[['Custo Unitário', 'Custo Total']].round(2)
    relatorio = relatorio.sort_values(['Linha', 'Subgrupo', 'Grade'])
    
    return relatorio

def gerar_relatorio_geral(estoque, vendas, ecommerce):
    """Gera relatório geral (como era antes)"""
    print("Gerando relatório geral...")
    
    # Agrupamento geral
    relatorio = (estoque
        .groupby(['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE'], observed=True)
        .agg({'CUSTO_TOTAL_ITEM': 'sum', 'ESTOQUE': 'sum'})
        .reset_index()
        .assign(CUSTO_UNITARIO=lambda x: x['CUSTO_TOTAL_ITEM'] / x['ESTOQUE'])
    )
    
    # Integra vendas se disponível
    if vendas is not None or ecommerce is not None:
        print("Integrando vendas gerais...")
        vendas_mensais = calcular_vendas_mensais_geral(vendas, ecommerce)
        relatorio = relatorio.merge(vendas_mensais, on=['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE'], how='left')
        
        # Preenche NaN
        for col in COLUNAS_MESES:
            if col in relatorio.columns:
                relatorio[col] = relatorio[col].fillna(0)
        
        # Métricas
        print("Calculando métricas...")
        meses_calc = min(3, len(COLUNAS_MESES))
        ultimos_meses = COLUNAS_MESES[-meses_calc:] if meses_calc > 0 else []
        
        if ultimos_meses:
            relatorio['Projeção Mensal'] = relatorio[ultimos_meses].mean(axis=1).round(0)
            meses_restantes = max(12 - MES_ATUAL + 1, 0)
            relatorio['Projeção Ano'] = (relatorio['Projeção Mensal'] * meses_restantes).round(0)
            relatorio['Estoque Final Mês'] = (relatorio['ESTOQUE'] - relatorio['Projeção Mensal']).round(0)
            relatorio['Estoque Final Ano'] = (relatorio['ESTOQUE'] - relatorio['Projeção Ano']).round(0)
            relatorio['Dias de Estoque'] = (relatorio['ESTOQUE'] / relatorio['Projeção Mensal'].replace(0, 1) * 30).round(1)
            relatorio.loc[relatorio['Projeção Mensal'] == 0, 'Dias de Estoque'] = 999
    
    # Renomeia e ordena
    relatorio = (relatorio
        .rename(columns={
            'LINHA': 'Linha',
            'SUBGRUPO_PRODUTO': 'Subgrupo',
            'GRADE': 'Grade',
            'CUSTO_UNITARIO': 'Custo Unitário',
            'CUSTO_TOTAL_ITEM': 'Custo Total',
            'ESTOQUE': 'Estoque Total'
        })
    )
    
    # Ordem das colunas
    cols_ordem = ['Linha', 'Subgrupo', 'Grade', 'Custo Unitário', 'Custo Total', 'Estoque Total']
    if vendas is not None or ecommerce is not None:
        cols_ordem.extend(COLUNAS_MESES + ['Projeção Mensal', 'Projeção Ano', 'Estoque Final Mês', 'Estoque Final Ano', 'Dias de Estoque'])
    
    relatorio = relatorio[[c for c in cols_ordem if c in relatorio.columns]]
    relatorio[['Custo Unitário', 'Custo Total']] = relatorio[['Custo Unitário', 'Custo Total']].round(2)
    relatorio = relatorio.sort_values(['Linha', 'Subgrupo', 'Grade'])
    
    return relatorio

def calcular_vendas_mensais_geral(vendas, ecommerce):
    """Calcula vendas mensais gerais integrando ecommerce"""
    print("Calculando vendas mensais gerais...")
    
    # Processa vendas tratadas
    vendas_tratadas = pd.DataFrame()
    if vendas is not None:
        vendas_tratadas = (vendas
            .assign(
                DATA_VENDA=lambda x: pd.to_datetime(x['DATA_VENDA'], errors='coerce'),
                LINHA=lambda x: x['LINHA'].str.strip(),
                SUBGRUPO_PRODUTO=lambda x: x['SUBGRUPO_PRODUTO'].str.strip(),
                GRADE=lambda x: x['GRADE'].str.strip(),
                QTDE=lambda x: pd.to_numeric(x['QTDE'], errors='coerce').fillna(0)
            )
            .dropna(subset=['DATA_VENDA'])
            .query(f'DATA_VENDA.dt.year == {ANO_ATUAL}')
            .assign(MES=lambda x: x['DATA_VENDA'].dt.month)
        )[['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE', 'MES', 'QTDE']]
    
    # Processa ecommerce se disponível
    vendas_ecommerce = pd.DataFrame()
    if ecommerce is not None:
        vendas_ecommerce = (ecommerce
            .assign(
                DATA_VENDA=lambda x: pd.to_datetime(x['DATA_SAIDA'], errors='coerce'),
                LINHA=lambda x: x['LINHA'].str.strip(),
                SUBGRUPO_PRODUTO=lambda x: x['SUBGRUPO_PRODUTO'].str.strip(),
                GRADE=lambda x: x['GRADE'].str.strip(),
                QTDE=lambda x: pd.to_numeric(x['QTDE'], errors='coerce').fillna(0)
            )
            .dropna(subset=['DATA_VENDA'])
            .query(f'DATA_VENDA.dt.year == {ANO_ATUAL} and QTDE > 0')
            .assign(MES=lambda x: x['DATA_VENDA'].dt.month)
        )[['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE', 'MES', 'QTDE']]
    
    # Combina dados
    vendas_combinadas = pd.concat([vendas_tratadas, vendas_ecommerce], ignore_index=True)
    
    if vendas_combinadas.empty:
        return pd.DataFrame()
    
    # Pivot direto
    vendas_pivot = (vendas_combinadas
        .groupby(['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE', 'MES'], observed=True)['QTDE']
        .sum()
        .unstack(fill_value=0)
        .reset_index()
        .rename(columns={i: COLUNAS_MESES[i-1] for i in range(1, MES_ATUAL+1) if i in range(1, 13)})
    )
    
    # Adiciona meses faltantes
    for col in COLUNAS_MESES:
        if col not in vendas_pivot.columns:
            vendas_pivot[col] = 0
    
    return vendas_pivot

def gerar_relatorio():
    """Processa e gera relatório com aba geral + abas por filial"""
    print("="*60)
    print("RELATÓRIO DE ESTOQUE GERAL + FILIAIS - SCARFME")
    print("="*60)

    tipo_analise = obter_tipo_analise()
    modo_filtros = None
    filtros_item = {}
    if tipo_analise == "filtrada":
        modo_filtros = obter_modo_combinacao_filtros()
        filtros_item = obter_filtros_item()
    
    # Carrega dados
    print("\nCarregando dados...")
    try:
        estoque = pd.read_csv('data/estoque_tratados.csv', sep=';', encoding='utf-8',
                             dtype={'PRODUTO': str, 'COR_PRODUTO': str}, low_memory=False)
        
        vendas = None
        ecommerce = None
        
        if os.path.exists('data/vendas_tratadas.csv'):
            vendas = pd.read_csv('data/vendas_tratadas.csv', sep=';', encoding='utf-8',
                               dtype={'PRODUTO': str, 'COR_PRODUTO': str}, low_memory=False)
            print(f"✓ Vendas: {len(vendas):,}")
        
        if os.path.exists('data/ecommerce.csv'):
            ecommerce = pd.read_csv('data/ecommerce.csv', sep=';', encoding='utf-8',
                                   dtype={'PRODUTO': str, 'COR_PRODUTO': str}, low_memory=False)
            print(f"✓ Ecommerce: {len(ecommerce):,}")
    except Exception as e:
        print(f"✗ Erro: {e}")
        return
    
    print(f"✓ Estoque: {len(estoque):,}")
    
    # Valida colunas
    cols_req = ['LINHA', 'SUBGRUPO_PRODUTO', 'GRADE', 'ESTOQUE', 'CUSTO_REPOSICAO1', 'FILIAL']
    if missing := [c for c in cols_req if c not in estoque.columns]:
        print(f"✗ Colunas faltando: {', '.join(missing)}")
        return
    
    # Pipeline de processamento base
    print("Processando dados base...")
    estoque = (estoque
        .assign(
            ESTOQUE=lambda x: pd.to_numeric(x['ESTOQUE'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0),
            CUSTO_REPOSICAO1=lambda x: pd.to_numeric(x['CUSTO_REPOSICAO1'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0),
            LINHA=lambda x: x['LINHA'].str.strip(),
            SUBGRUPO_PRODUTO=lambda x: x['SUBGRUPO_PRODUTO'].str.strip(),
            GRADE=lambda x: x['GRADE'].str.strip(),
            FILIAL=lambda x: x['FILIAL'].str.strip()
        )
        .query('ESTOQUE > 0 and CUSTO_REPOSICAO1 >= 0 and LINHA != "ELETRONICOS"')
        .assign(CUSTO_TOTAL_ITEM=lambda x: x['ESTOQUE'] * x['CUSTO_REPOSICAO1'])
    )

    # Se análise filtrada, aplicar filtros nos dados (estoque + vendas + ecommerce)
    if tipo_analise == "filtrada" and filtros_item:
        print("\nAplicando filtros (modo: %s)..." % modo_filtros)

        # Estoque não tem COLECAO -> enriquecer via produtos_tratados se precisar filtrar por coleção
        if 'colecao' in filtros_item and 'COLECAO' not in estoque.columns and os.path.exists('data/produtos_tratados.csv'):
            try:
                produtos = pd.read_csv('data/produtos_tratados.csv', sep=';', encoding='utf-8',
                                      dtype={'PRODUTO': str}, low_memory=False)
                if 'PRODUTO' in produtos.columns and 'COLECAO' in produtos.columns:
                    produtos = produtos[['PRODUTO', 'COLECAO']].copy()
                    produtos['PRODUTO'] = produtos['PRODUTO'].astype(str).str.strip()
                    produtos['COLECAO'] = produtos['COLECAO'].astype(str).str.strip().str.upper()
                    estoque['PRODUTO'] = estoque['PRODUTO'].astype(str).str.strip()
                    estoque = estoque.merge(produtos, on='PRODUTO', how='left')
                    print("✓ COLECAO adicionada no estoque via produtos_tratados.csv")
            except Exception as e:
                print(f"⚠️ Não foi possível carregar produtos_tratados.csv para COLECAO: {e}")

        estoque = aplicar_filtros_dinamicos(estoque, filtros_item, modo_filtros)
        print(f"✓ Estoque após filtros: {len(estoque):,}")

        if vendas is not None and not vendas.empty:
            vendas = aplicar_filtros_dinamicos(vendas, filtros_item, modo_filtros)
            print(f"✓ Vendas após filtros: {len(vendas):,}")

        if ecommerce is not None and not ecommerce.empty:
            ecommerce = aplicar_filtros_dinamicos(ecommerce, filtros_item, modo_filtros)
            print(f"✓ Ecommerce após filtros: {len(ecommerce):,}")
    
    # Gera relatório geral
    relatorio_geral = gerar_relatorio_geral(estoque, vendas, ecommerce)
    print(f"✓ Geral: {len(relatorio_geral)} categorias")
    
    # Processa cada filial
    print("\nProcessando filiais...")
    relatorios_filiais = {}
    
    for filial_original, filial_nome in FILIAIS.items():
        relatorio = processar_filial(estoque, vendas, ecommerce, filial_nome, filial_original)
        if relatorio is not None:
            relatorios_filiais[filial_nome] = relatorio
            print(f"    ✓ {filial_nome}: {len(relatorio)} categorias")
    
    # Salva Excel com múltiplas abas
    os.makedirs('relatorios', exist_ok=True)
    arquivo = gerar_nome_arquivo(tipo_analise, modo_filtros, filtros_item)
    
    try:
        print(f"\nSalvando: {arquivo}")
        with pd.ExcelWriter(arquivo, engine='openpyxl', mode='w') as writer:
            # Primeira aba: Geral
            relatorio_geral.to_excel(writer, sheet_name='Geral', index=False)
            ws = writer.sheets['Geral']
            
            # Formata colunas da aba geral
            for idx, col in enumerate(ws.columns, 1):
                ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value)) for c in col) + 2, 50)
                if idx in [4, 5]:  # Custo Unitário e Total
                    for cell in list(col)[1:]:
                        cell.number_format = 'R$ #,##0.00'
            
            aplicar_formatacao_cores(ws, relatorio_geral, "Estoque Total")
            
            # Abas das filiais
            for filial_nome, relatorio in relatorios_filiais.items():
                relatorio.to_excel(writer, sheet_name=filial_nome, index=False)
                ws = writer.sheets[filial_nome]
                
                # Formata colunas
                for idx, col in enumerate(ws.columns, 1):
                    ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value)) for c in col) + 2, 50)
                    if idx in [4, 5]:  # Custo Unitário e Total
                        for cell in list(col)[1:]:
                            cell.number_format = 'R$ #,##0.00'
                
                aplicar_formatacao_cores(ws, relatorio, f"Estoque {filial_nome}")
        
        print(f"✓ Salvo: {arquivo}")
    except PermissionError:
        base, ext = os.path.splitext(arquivo)
        arquivo = f"{base}_{datetime.now():%Y%m%d_%H%M%S}{ext}"
        print(f"⚠️ Arquivo aberto, criando: {arquivo}")
        with pd.ExcelWriter(arquivo, engine='openpyxl') as writer:
            # Primeira aba: Geral
            relatorio_geral.to_excel(writer, sheet_name='Geral', index=False)
            ws = writer.sheets['Geral']
            aplicar_formatacao_cores(ws, relatorio_geral, "Estoque Total")
            
            # Abas das filiais
            for filial_nome, relatorio in relatorios_filiais.items():
                relatorio.to_excel(writer, sheet_name=filial_nome, index=False)
                ws = writer.sheets[filial_nome]
                aplicar_formatacao_cores(ws, relatorio, f"Estoque {filial_nome}")
    
    # Resumo
    print(f"\n✓ Abas geradas: {1 + len(relatorios_filiais)} (Geral + {len(relatorios_filiais)} filiais)")
    print(f"✓ Geral: {len(relatorio_geral)} categorias, R$ {relatorio_geral['Custo Total'].sum():,.2f}")
    
    for filial_nome, relatorio in relatorios_filiais.items():
        print(f"  - {filial_nome}: {len(relatorio)} categorias, R$ {relatorio['Custo Total'].sum():,.2f}")
    
    print("\n" + "="*60)
    print("CONCLUÍDO!")
    print("="*60)

if __name__ == "__main__":
    try:
        gerar_relatorio()
    except Exception as e:
        print(f"\n✗ Erro: {e}")
        import traceback
        traceback.print_exc()